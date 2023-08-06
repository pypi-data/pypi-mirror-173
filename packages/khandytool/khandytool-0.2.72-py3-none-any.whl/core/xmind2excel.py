import os
from openpyxl import Workbook
from xmindparser import xmind_to_dict
import traceback


def xmind2List(file_path):
    try:
        xmind_origin = xmind_to_dict(file_path)
        source=xmind_origin[0]['topic']['topics']
        # print("*"*50+f'\n{source}\n'+"*"*50)
        temp_list=[]
        all_list=[]

        for one_m in source:
            temp_list.append(one_m['title'])
            for one_s in one_m['topics']:
                temp_list.append(one_s['title'])
                for one_t in one_s['topics']:
                    temp_list.append(one_t['title'])
                    for one_p in one_t['topics']:
                        temp_list.append(one_p['title'])
                        for one_e in one_p['topics']:
                            temp_list.append(one_e['title'])
                    all_list.append(temp_list)
                    temp_list=[]
    except Exception as e:
        print(traceback.format_exc())
    return all_list

def makeCase(xmindPath,excelName):
    try:
        wb = Workbook()
        ws = wb.active
        row=1
        title=['模块','用例组','用例标题','用例步骤','步骤期望']
        for index,value in enumerate(title):
            ws.cell(row,index+1,value=value)  
        row=row+1
        for one_row in xmind2List(xmindPath):
            if len(one_row)==5:
                # print(one_row)
                for index,value in enumerate(one_row):
                    ws.cell(row,index+1,value=value)
            elif len(one_row)==4:
                one_row.insert(0,'')
                # print(one_row)
                for index,value in enumerate(one_row):
                    ws.cell(row,index+1,value=value)
            elif len(one_row)==3:
                one_row.insert(0,'')
                one_row.insert(1,'')
                # print(one_row)
                for index,value in enumerate(one_row):
                    ws.cell(row,index+1,value=value)
            row=row+1
        wb.save(excelName)   
    except Exception as e:
        print(traceback.format_exc())
                    
if __name__=='__main__':
    makeCase(xmindPath="./test/测试xmind.xmind",excelName='./test/测试.xlsx')
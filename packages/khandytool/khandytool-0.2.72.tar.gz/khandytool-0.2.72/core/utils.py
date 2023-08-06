#coding:utf-8
import jmespath
import hashlib
from faker import Faker
import time
from datetime import datetime
import pymysql
from loguru import logger
from functools import wraps
import os
import json
import re
from xml.dom.minidom import parse
import xmltodict
from jsonpath import jsonpath
import wget
import traceback
import inspect
import ctypes


def mkdirAfterCheck(path):
    if not os.path.exists(path):
        os.mkdir(path)
    
def mkfileAfterCheck(path,fileName):
    mkdirAfterCheck(path)
    if not os.path.exists(path+os.sep+fileName):
        open(path+os.sep+fileName,'w',encoding='utf-8')
            
    


def timeit(func):
    '''
        func：funtion name which need calculate time
    '''
    @wraps(func)
    def calculate(*args,**kwargs):
        start=datetime.now()
        func(*args,**kwargs)
        end=datetime.now()
        logger.info(f"{func.__name__} costs {(end-start).seconds}")
        return calculate


def addException(func):
    @wraps(func)
    def innerFun(*args,**kwargs):
        try:
            res=None
            res=func(*args,**kwargs)
        except Exception as e:
            print(e)
            print('Exception found, detail below:')
            print(traceback.format_exc())
        return res
    return innerFun


def singleton(cls):
    instance = {}
    @wraps(cls)
    def get_insance(*args, **kwargs):
        if cls not in instance:
            instance[cls] = cls(*args, **kwargs)
        return instance[cls]
    return get_insance



def jsonPrettyOutput(json_string):
    try:
        parsed_json = json.loads(json_string)
        formatted_json = json.dumps(parsed_json, indent = 4,sort_keys=True)
        return formatted_json
    except Exception as e:
        return e






@addException
def email_sender(host="smtp.163.com",sender="xxx@163.com",auth_code='xxx',reciever=['xxx','xxx'],content="email content"):
    import smtplib
    import email

    # 负责将多个对象集合起来
    from email.mime.multipart import MIMEMultipart
    from email.header import Header

    # SMTP服务器,这里使用163邮箱
    mail_host = host
    # 发件人邮箱
    mail_sender = sender
    # 邮箱授权码,注意这里不是邮箱密码,如何获取邮箱授权码,请看本文最后教程
    mail_license = auth_code
    # 收件人邮箱，可以为多个收件人
    mail_receivers = reciever

    mm = MIMEMultipart('related')
    # 邮件正文内容
    body_content = content
    # 构造文本,参数1：正文内容，参数2：文本格式，参数3：编码方式
    message_text = MIMEText(body_content,"plain","utf-8")
    # 向MIMEMultipart对象中添加文本对象
    mm.attach(message_text)

    # 创建SMTP对象
    stp = smtplib.SMTP()
    # 设置发件人邮箱的域名和端口，端口地址为25
    stp.connect(mail_host, 25)  
    # set_debuglevel(1)可以打印出和SMTP服务器交互的所有信息
    stp.set_debuglevel(1)
    # 登录邮箱，传递参数1：邮箱地址，参数2：邮箱授权码
    stp.login(mail_sender,mail_license)
    # 发送邮件，传递参数1：发件人邮箱地址，参数2：收件人邮箱地址，参数3：把邮件内容格式改为str
    stp.sendmail(mail_sender, mail_receivers, mm.as_string())
    print("邮件发送成功")
    # 关闭SMTP对象
    stp.quit()



def download2Dest(url,destDir,fileName):
    try:
        if not os.path.exists(destDir):
            os.mkdir(destDir)
        wget.download(url,out=destDir+os.sep+fileName)
        return True
    except Exception as e:
        logger.error(e)
        return False
    return True


class strFilter(object):
    def __init__(self,oriStr,targetStr,flag="",pattern="",retFlag="yes"):
        self.oriStr=oriStr
        self.targetStr=targetStr
        self.flag=flag
        self.pattern=pattern
        self.retFlag=retFlag
    def falgValidate(self):
        if flag=="":
            self.filterContain()
        elif flag=="json":
            self.filterJson()
        elif flag=="regx":
            self.filterRegx()
        else:
            logger.error('your choose wrong flter flag')
    def filterContain(self):
        if self.retFlag.lower()=="no":
            if self.targetStr in self.oriStr:
                return True
            else:
                return False
        else:
            if self.targetStr in self.oriStr:
                return self.oriStr
            else:
                return None
    def filterJson(self):
        if self.retFlag.lower()=="no":
            if self.targetStr in self.oriStr:
                return True
            else:
                return False
        else:
            if self.targetStr in self.oriStr:
                return self.oriStr
            else:
                return None




def findBy(oriStr,flag="",pattern=""):
    if flag=="json":
        if isJson(str(oriStr)):
            try:
                allStr = json.loads(oriStr)
                flag=jmespath.search(pattern,allStr)
                if flag:
                    return flag
                else:
                    return False
            except Exception as e:
                logger.error('json 解析失败')
                return False  
    elif flag=="regx":
        try:
            allStr = oriStr
            flag=re.compile(pattern)
            res=flag.findall(allStr)
            targetStr="".join(res)
            if targetStr:
                return targetStr
            else:
                return False
        except Exception as e:
            logger.error('regx 解析失败')
            return False
    else:
        logger.error(f'无此{flag}解析方式')



def jsonFilter(oriStr,flag,pattern,key):
    if isJson(str(oriStr)):
        try:
            allStr = json.loads(oriStr)
            flag=jmespath.search(pattern,allStr)
            if flag in key:
                return True, allStr
        except Exception as e:
            logger.error('json 解析失败')
            logger.error(f'{allStr}')
            return False, allStr
    else:
        logger.error(f"{allStr} isn't a string.")
        return False, allStr

def regxFilter(oriStr,flag,pattern,key):
    try:
        allStr = oriStr
        flag=re.compile(pattern)
        res=flag.findall(allStr)
        if key in "".join(res):
            return True, allStr
        return False, allStr
    except Exception as e:
        logger.error('regx 解析失败')
        logger.error(f'{allStr}') 
        return False, allStr

def containFilter(oriStr,key):
    try:
        if key in oriStr:
            return True, oriStr
        else:
            return False, oriStr
    except Exception as e:
        logger.error('contain 错误')
        return False, oriStr

def checkStr(oriStr,flag,pattern,key):
        if flag=="json":
            if jsonFilter(oriStr, flag, pattern, key)[0]!=False:
                return True
            else:
                return False
        elif flag=="regx":
            if regxFilter(oriStr, flag, pattern, key)[0]!=False:
                return True
            else:
                return False
        elif flag=="contain":
            if containFilter(oriStr, key)[0]!=False:
                return True
            else:
                return False

def getFromList(input_list=[],list_index=0):
    if len(input_list)==1:
        return input_list[0]
    else:
        return "wrong num of input_list"


#多维数组转单维数组
singleList = []
def multi2OneList(mylist):
    '''
    mylist: list should be larger than 1 demi
    '''
    global singleList
    for one in mylist:
        if isinstance(one, list):
            multi2OneList(one)
        else:
            singleList.append(one)
    return singleList

#简单版本

def multi2Single(*args):
    ret = []
    for one in args:
        ret = ret+one
    return ret


#判断是否为json串
def isJson(oriStr):
    if isinstance(oriStr,str):
        try:
            targetStr=json.loads(oriStr)
        except ValueError as e:
            return False
        return True
    else:
        logger.error(f'{oriStr} not a string')
        return False


#获取数组维数
num = 1
def getDimons(mylist):
    global num
    for one in mylist:
        # print(one)
        if isinstance(one, list):
            num = num+1
            getDimons(one)
    return num



dbInfo={}
class mysql_exe(object):
    def __init__(self,host="",user="",password="",port="",db="",charset=""):
        if host=="":
            self.conn = pymysql.connect(host=dbInfo['host'],
                                        user=dbInfo['username'],
                                        password=dbInfo['password'],
                                        port=dbInfo['port'],
                                        db=dbInfo['dbname'],
                                        charset=dbInfo['charset'])
            self.cur = self.conn.cursor()
        self.conn = pymysql.connect(host=host,user=user,password=password,port=int(port),db=db,charset=charset)
        self.cur = self.conn.cursor()

    def __enter__(self):
        return self

    def run(self, sql):
        try:
            self.cur.execute(sql)
            self.conn.commit()
        except Exception as e:
            self.conn.rollback()
            logger.error(e)

    def run_with_print(self, sql):
        try:
            self.cur.execute(sql)
            self.conn.commit()
            data = self.cur.fetchall()
            for one in data:
                logger.info(f'database return is : {one}')
        except Exception as e:
            self.conn.rollback()
            logger.error(e)

    def run_with_return(self, sql):
        try:
            self.cur.execute(sql)
            self.conn.commit()
            data = self.cur.fetchall()
            return data
        except Exception as e:
            self.conn.rollback()
            logger.error(e)

    def __exit__(self, exc_type, exc_val, exc_tb):
        self.cur.close()
        self.conn.close()


@addException
def getSha1Password(passwdStr,key="terminus2021"):
    '''

    :param passwdStr: password
    :param key: salt
    :return: sha1 encrypt
    '''
    sh=hashlib.sha1(bytes(key+passwdStr,encoding='utf-8')).hexdigest()
    return sh

@addException
def getData(keyStr,businessData):
    '''

    :param keyStr: search patten
    :param businessData: total data
    :return: key or key list
    '''
    flag=jmespath.search(keyStr,businessData)
    # print(flag)
    if not flag:
        return 'not found'
    return flag



class CFacker():
    '''
    generate fake data
    '''
    city_suffix = '：市，县'
    country = '：国家'
    country_code = '：国家编码'
    district = '：区'
    geo_coordinate = '：地理坐标'
    latitude = '：地理坐标(纬度)'
    longitude = '：地理坐标(经度)'
    postcode = '：邮编'
    province = '：省份 (zh_TW没有此方法)'
    address = '：详细地址'
    street_address = '：街道地址'
    street_name = '：街道名'
    street_suffix = '：街、路'
    ssn = '：生成身份证号'
    bs = '：随机公司服务名'
    company = '：随机公司名（长）'
    company_prefix = '：随机公司名（短）'
    company_suffix = '：公司性质'
    credit_card_expire = '：随机信用卡到期日'
    credit_card_full = '：生成完整信用卡信息'
    credit_card_number = '：信用卡号'
    credit_card_provider = '：信用卡类型'
    credit_card_security_code = '：信用卡安全码'
    job = '：随机职位'
    first_name = '：名'
    first_name_female = '：女性名'
    first_name_male = '：男性名'
    first_romanized_name = '：罗马名'
    last_name = '：姓'
    last_name_female = '：女姓'
    last_name_male = '：男姓'
    last_romanized_name = '：随机'
    name = '：随机生成全名'
    name_female = '：男性全名'
    name_male = '：女性全名'
    romanized_name = '：罗马名'
    msisdn = '：移动台国际用户识别码，即移动用户的ISDN号码'
    phone_number = '：随机生成手机号'
    phonenumber_prefix = '：随机生成手机号段'
    ascii_company_email = '：随机ASCII公司邮箱名'
    ascii_email = '：随机ASCII邮箱'
    ascii_free_email = '：二进制免费邮件'
    ascii_safe_email = '：二进制安全邮件'
    company_email = '：公司邮件'
    email = '：电子邮件'
    free_email = '：免费电子邮件'
    free_email_domain = '：免费电子邮件域名'
    safe_email = '：安全邮箱'
    domain_name = '：生成域名'
    domain_word = '：域词(即，不包含后缀)'
    ipv4 = '：随机IP4地址'
    ipv6 = '：随机IP6地址'
    mac_address = '：随机MAC地址'
    tld = '：网址域名后缀(.com,.net.cn,等等，不包括.)'
    uri = '：随机URI地址'
    uri_extension = '：网址文件后缀'
    uri_page = '：网址文件（不包含后缀）'
    uri_path = '：网址文件路径（不包含文件名）'
    url = '：随机URL地址'
    user_name = '：随机用户名'
    image_url = '：随机URL地址'
    chrome = '：随机生成Chrome的浏览器user_agent信息'
    firefox = '：随机生成FireFox的浏览器user_agent信息'
    internet_explorer = '：随机生成IE的浏览器user_agent信息'
    opera = '：随机生成Opera的浏览器user_agent信息'
    safari = '：随机生成Safari的浏览器user_agent信息'
    linux_platform_token = '：随机Linux信息'
    user_agent = '：随机user_agent信息'
    file_extension = '：随机文件扩展名'
    file_name = '：随机文件名（包含扩展名，不包含路径）'
    file_path = '：随机文件路径（包含文件名，扩展名）'
    mime_type = '：随机mime Type'
    numerify = '：三位随机数字'
    random_digit = '：0~9随机数'
    random_digit_not_null = '：1~9的随机数'
    random_int = '：随机数字，默认0~9999，可以通过设置min,max来设置'
    random_number = '：随机数字，参数digits设置生成的数字位数'
    pyfloat = '：left_digits=5 #生成的整数位数, right_digits=2 #生成的小数位数, positive=True #是否只有正数'
    pyint = '：随机Int数字（参考random_int=参数）'
    pydecimal = '：随机Decimal数字（参考pyfloat参数）'
    pystr = '：随机字符串'
    random_element = '：随机字母'
    random_letter = '：随机字母'
    paragraph = '：随机生成一个段落'
    paragraphs = '：随机生成多个段落，通过参数nb来控制段落数，返回数组'
    sentence = '：随机生成一句话'
    sentences = '：随机生成多句话，与段落类似'
    text = '：随机生成一篇文章（不要幻想着人工智能了，至今没完全看懂一句话是什么意思）'
    word = '：随机生成词语'
    words = '：随机生成多个词语，用法与段落，句子，类似'
    binary = '：随机生成二进制编码'
    boolean = '：True/False'
    language_code = '：随机生成两位语言编码'
    locale = '：随机生成语言/国际 信息'
    md5 = '：随机生成MD5'
    null_boolean = '：NULL/True/False'
    password = '：随机生成密码,可选参数：length：密码长度；special_chars：是否能使用特殊字符；digits：是否包含数字；upper_case：是否包含大写字母；lower_case：是否包含小写字母'
    sha1 = '：随机SHA1'
    sha256 = '：随机SHA256'
    uuid4 = '：随机UUID'
    am_pm = '：AM/PM'
    century = '：随机世纪'
    date = '：随机日期'
    date_between = '：随机生成指定范围内日期，参数：start_date，end_date取值：具体日期或者today,-30d,-30y类似'
    date_between_dates = '：随机生成指定范围内日期，用法同上'
    date_object = '：随机生产从1970-1-1到指定日期的随机日期。'
    date_this_month = '：当月'
    date_this_year = '：当年'
    date_time = '：随机生成指定时间（1970年1月1日至今）'
    date_time_ad = '：生成公元1年到现在的随机时间'
    date_time_between = '：用法同dates'
    future_date = '：未来日期'
    future_datetime = '：未来时间'
    month = '：随机月份'
    month_name = '：随机月份（英文）'
    past_date = '：随机生成已经过去的日期'
    past_datetime = '：随机生成已经过去的时间'
    time = '：随机24小时时间'
    timedelta = '：随机获取时间差'
    time_object = '：随机24小时时间，time对象'
    time_series = '：随机TimeSeries对象'
    timezone = '：随机时区'
    unix_time = '：随机Unix时间'
    year = '：随机年份'
    profile = '：随机生成档案信息'
    simple_profile = '：随机生成简单档案信息'
    currency_code = '：货币编码'
    color_name = '：随机颜色名'
    hex_color = '：随机HEX颜色'
    rgb_color = '：随机RGB颜色'
    safe_color_name = '：随机安全色名'
    safe_hex_color = '：随机安全HEX颜色'
    isbn10 = '：随机ISBN（10位）'
    isbn13 = '：随机ISBN（13位）'
    lexify = '：替换所有问号（“？”）带有随机字母的事件。'

    @classmethod
    def get_it(cls, name, *args, local='zh_CN'):
        fake = Faker(locale=local)
        str1=""
        for k,v in cls.__dict__.items():
            if name==k:
                # print(v)
                if len(args) != 0:
                    str1 = "fake.{}({})".format(k, *args)
                else:
                    str1 = "fake.{}()".format(k)
                # print(str1)
                return eval(str1)


def getTimeStamp():
    '''
    :return: get 13 timestamp
    '''
    return round(time.time() * 1000)

def formatStr2Timestamp(timeStr: str, format_style='%Y-%m-%d %H:%M:%S.%f'):
    """
    timeStr:"2022-01-01 08:08:08.999"
    """
    # time_array = time.strptime(timeStr, format_type)
    # return int(time.mktime(time_array)) * 1000

    datetime_obj = datetime.strptime(timeStr, format_style)
    ret_stamp = int(time.mktime(datetime_obj.timetuple()) * 1000.0 + datetime_obj.microsecond / 1000.0)
    return ret_stamp


def timeStampStr2FormatTime(timestamp: int, format_style='%Y-%m-%d %H:%M:%S.%f'):
    """"
    timestampStr:1648741273998
    """
    # return time.strftime(format_type,time.localtime(timestampStr))

    timestamp = float(timestamp) / 1000
    # ret_datetime = datetime.utcfromtimestamp(timestamp).strftime(format_style)
    ret_datetime = datetime.fromtimestamp(timestamp).strftime(format_style)
    # datetime.fromtimestamp(t)
    return ret_datetime


def timestampsCompare(bigOne: int, smarllOne: int):
    """
    bigOne:1648792986000
    smarllOne:1648792990000
    """
    big = datetime.utcfromtimestamp(float(bigOne) / 1000).strftime('%Y-%m-%d %H:%M:%S.%f')
    bigt = datetime.strptime(big, '%Y-%m-%d %H:%M:%S.%f')
    # print(bigt)
    small = datetime.utcfromtimestamp(float(smarllOne) / 1000).strftime('%Y-%m-%d %H:%M:%S.%f')
    smallt = datetime.strptime(small, '%Y-%m-%d %H:%M:%S.%f')
    # print(smallt)
    dalta = bigt - smallt
    # return dalta.milliseconds
    return dalta.total_seconds()


def timeComparedBySeconds(starter,ender):
    '''
    #starter=datetime.now()
    #time.sleep(3)
    #ender=datetime.now()
    '''
    num=(ender-starter).seconds
    return num


def getDateTime(format="%Y-%m-%d %H:%M:%S.%f"):
    '''

    :param format: time format
    :return: format time
    '''
    return datetime.strftime(datetime.now(), format)

def jsonfile_to_obj(json_file_path):
    # import json
    # from collections import namedtuple
    # json_file=open(file=json_file_path, mode='r', encoding='utf-8')
    # pyobj=json.load(json_file,object_hook=lambda d: namedtuple('X',d.keys())(*d.values()))
    # return pyobj
    import json
    try:
        from types import SimpleNamespace as Namespace
    except ImportError:
        from argparse import Namespace
    json_obj=open(file=json_file_path, mode='r', encoding='utf-8')
    pyobj=json.load(json_obj,object_hook=lambda d: Namespace(**d))
    return pyobj


def json_to_obj(json_obj):
    # import json
    # from collections import namedtuple
    # json_file=open(file=json_file_path, mode='r', encoding='utf-8')
    # pyobj=json.load(json_file,object_hook=lambda d: namedtuple('X',d.keys())(*d.values()))
    # return pyobj
    import json
    try:
        from types import SimpleNamespace as Namespace
    except ImportError:
        from argparse import Namespace
    pyobj=json.loads(json_obj,object_hook=lambda d: Namespace(**d))
    return pyobj


from openpyxl import Workbook
from xmindparser import xmind_to_dict


def xmind2List(file_path):
    xmind_origin = xmind_to_dict(file_path)
    source=xmind_origin[0]['topic']['topics']
    # print("*"*50+f'\n{source}\n'+"*"*50)
    temp_list=[]
    all_list=[]

    for one_m in source:
        temp_list.append(one_m['title'])
        for one_s in one_m['topics']:
            temp_list.append(one_s['title'])
            for one_t in one_s['topics']:
                temp_list.append(one_t['title'])
                for one_p in one_t['topics']:
                    temp_list.append(one_p['title'])
                    for one_e in one_p['topics']:
                        temp_list.append(one_e['title'])
                all_list.append(temp_list)
                temp_list=[]

    return all_list

def makeCase(xmindPath,excelName):
    wb = Workbook()
    ws = wb.active
    row=1
    title=['模块','用例组','用例标题','用例步骤','步骤期望']
    for index,value in enumerate(title):
        ws.cell(row,index+1,value=value)  
    row=row+1
    for one_row in xmind2List(xmindPath):
        if len(one_row)==5:
            # print(one_row)
            for index,value in enumerate(one_row):
                ws.cell(row,index+1,value=value)
        elif len(one_row)==4:
            one_row.insert(0,'')
            # print(one_row)
            for index,value in enumerate(one_row):
                ws.cell(row,index+1,value=value)
        elif len(one_row)==3:
            one_row.insert(0,'')
            one_row.insert(1,'')
            # print(one_row)
            for index,value in enumerate(one_row):
                ws.cell(row,index+1,value=value)
        row=row+1
    wb.save(excelName)   


# def swagger2jmeter(url):
#     from swaggerjmx.convert import conversion
#     from swaggerjmx.settings import Settings as ST
#     import traceback
#     #  swagger_url
#     try:
#         ST.swagger_url = url
#         #  report_path
#         ST.report_path = 'jmx'
#         # 开始转换
#         conversion()
#         # print(os.path.abspath(os.path.dirname(__file__))+os.sep+'jmx'+os.sep)
#         upperDir=os.path.abspath(os.path.dirname(__file__))+os.sep+'jmx'+os.sep
#         for x,y,z in os.walk(os.path.abspath(os.path.dirname(__file__))+os.sep+'jmx'+os.sep):
#             return upperDir+"".join(z)
#     except Exception as e:
#         print(traceback.format_exc())
#         return "switch to jmeter script error"


def parseJmeterXml(filePath='/Users/oupeng/ssoHttp.jmx'):
    try:
        resList=[]
        with open(filePath,mode='r',encoding='utf8') as xf:
            doc=xmltodict.parse(xf.read())
            # print(doc['jmeterTestPlan']['hashTree']['hashTree']['Arguments']['collectionProp']['elementProp']['stringProp'])
            json_result=json.loads(json.dumps(doc,ensure_ascii=False))
            # print(json_result)
            args=jsonpath(json_result, '$..Arguments.collectionProp.elementProp[*].stringProp')
            # print(args)
            
            for arg in args:
                tempList=[]
                for one in arg:
                    if one['@name']=='Argument.name':
                        tempList.append(one['#text'])
                    elif one['@name']=='Argument.value':
                        if "${__P" in one['#text']:
                            tempList.append(one['#text'].strip('${__P(').strip(')}').split(',')[1])
                        else:
                            tempList.append(one['#text'])
                    elif one['@name']=='Argument.desc':
                        tempList.append(one['#text'])
                if len(tempList)>2:
                    # print(tempList)
                    if tempList not in resList:
                        resList.append(tempList)
            # print(resList)
        if len(resList)>0:
            return resList
        else:
            return False
    except Exception as e:
        print(e)
        

        
    # return paramList


def _async_raise(tid, exctype):
    """raises the exception, performs cleanup if needed"""
    tid = ctypes.c_long(tid)
    if not inspect.isclass(exctype):
        exctype = type(exctype)
    res = ctypes.pythonapi.PyThreadState_SetAsyncExc(tid, ctypes.py_object(exctype))
    if res == 0:
        raise ValueError("invalid thread id")
    elif res != 1:
        ctypes.pythonapi.PyThreadState_SetAsyncExc(tid, None)
        raise SystemError("PyThreadState_SetAsyncExc failed")


def stop_thread(thread):
    _async_raise(thread.ident, SystemExit)



if __name__=='__main__':
    # print(CFacker().get_it("month"))
    # print(swagger2jmeter("http://192.168.xxx.xxx:123456/space/v2/api-docs"))
    # print(formatStr2Timestamp("2022-04-01 08:08:08.345"))
    # print(timeStampStr2FormatTime(1648741273123))
    # timestampsCompare(1648741273123, 1648741259023)
    # print(parseJmeterXml())
    download2Dest("https://th.bing.com/th/id/OIP.hN3O2jHxpAjx7N_J9vmBYAHaBf?w=294&h=70&c=7&r=0&o=5&pid=1.7", os.path.expanduser('~')+os.sep+"abc.png")
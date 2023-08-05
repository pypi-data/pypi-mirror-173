import os
import openpyxl
import pickle
import numbers
import csv
import datetime
import json
from PyQt5.QtWidgets import QMainWindow, QToolBar, QFileDialog, QMessageBox, QListWidgetItem, QProgressDialog
from PyQt5.QtCore import Qt, QDate
from zhou_stattool.roomdata_window import Ui_MainWindow
from zhou_stattool.add_rule_window_main import AddRuleMainWindow


class RoomdataMainWindow(QMainWindow):

    DATETIME_FORMAT = '%Y/%m/%d %H:%M:%S'
    DATETIME_FORMAT_WITHOUT_SECOND = '%Y/%m/%d'

    def __init__(self):
        super(RoomdataMainWindow, self).__init__()
        self.ui = Ui_MainWindow()
        self.ui.setupUi(self)
        self.add_toolbar()
        self.bind_actions()
        self.init_vars()

    def add_toolbar(self):
        self.toolbar = QToolBar()
        self.addToolBar(Qt.TopToolBarArea, self.toolbar)
        self.toolbar.addAction(self.ui.actionopen_file)
        self.toolbar.addAction(self.ui.actionopen_dir)

    def bind_actions(self):
        self.ui.actionopen_dir.triggered.connect(self.open_dir)
        self.ui.actionopen_file.triggered.connect(self.open_file)
        self.ui.add_rule_btn.clicked.connect(self.add_rule)
        self.ui.remove_rule_btn.clicked.connect(self.remove_rule)
        self.ui.edit_rule_btn.clicked.connect(self.edit_rule)
        self.ui.open_rule_btn.clicked.connect(self.open_rule)
        self.ui.save_rule_btn.clicked.connect(self.save_rule)
        self.ui.rules_widget.itemClicked.connect(self.rule_item_clicked)
        self.ui.start_process_btn.clicked.connect(self.start_process)

    def init_vars(self):
        self.opened_files = list()
        self.table_header = list()
        self.rule = list()
        self.ui.start_process_btn.setEnabled(False)
        today = datetime.datetime.today()
        tomorrow = datetime.datetime(today.year, today.month, today.day + 1)
        self.ui.start_date_edit.setDate(today)
        self.ui.end_date_edit.setDate(tomorrow)
        self.ui.start_date_edit.setCalendarPopup(True)
        self.ui.end_date_edit.setCalendarPopup(True)
        self.ui.modeComboBox.addItems(['总表模式', '分表模式'])

    def open_dir(self, checked):
        file_dialog = QFileDialog(self)
        file_dialog.setFileMode(QFileDialog.DirectoryOnly)
        dir_name = file_dialog.getExistingDirectory()
        file_names = os.listdir(dir_name)
        for file_name in file_names:
            if not file_name.endswith('xlsx'):
                QMessageBox.critical(self, '错误', '文件夹中包含非表格文件', QMessageBox.Yes)
                return
        file_names = [os.path.join(dir_name, item) for item in file_names]
        self.opened_files = file_names
        self.show_opened_files()


    def open_file(self, checked):
        file_dialog = QFileDialog(self)
        file_dialog.setFileMode(QFileDialog.ExistingFiles)
        file_names = file_dialog.getOpenFileNames(filter='表格文件(*.xlsx)')[0]
        if file_names and len(file_names) > 0:
            self.opened_files = file_names
            self.show_opened_files()
        else:
            QMessageBox.critical(self, '错误', '请选择正确文件', QMessageBox.Yes)

    def check_file_headers(self):
        if not self.opened_files:
            return True, []
        wb = openpyxl.load_workbook(self.opened_files[0])
        sheet = wb['Sheet1']
        num_columns = sheet.max_column
        header = [sheet.cell(1, i).value.strip() for i in range(1, num_columns + 1)]
        for i in range(1, len(self.opened_files)):
            wb = openpyxl.load_workbook(self.opened_files[i])
            sheet = wb['Sheet1']
            if sheet.max_column != num_columns:
                return False, []
            for col in range(1, num_columns + 1):
                val = sheet.cell(1, col).value.strip()
                if val != header[col - 1]:
                    return False, []
        return True, header

    def show_opened_files(self):
        ret, header = self.check_file_headers()
        if ret and header:
            self.ui.opened_files_widget.clear()
            for file_name in self.opened_files:
                show_name = os.path.basename(file_name)
                self.ui.opened_files_widget.addItem(show_name)
            self.table_header = header
            self.ui.excel_header_widget.clear()
            for item in header:
                self.ui.excel_header_widget.addItem(item)
        else:
            self.opened_files.clear()
            self.ui.opened_files_widget.clear()
            QMessageBox.critical(self, '错误', '选择文件的表格头不一致', QMessageBox.Yes)

    def add_rule(self):
        if self.opened_files:
            self.add_rule_window = AddRuleMainWindow(self.table_header)
            self.add_rule_window.setWindowModality(Qt.ApplicationModal)
            ret = self.add_rule_window.exec_()
            if ret:
                name, indexes = self.add_rule_window.get_result()
                indexes = [item.row() for item in indexes]
                if not name or not indexes:
                    QMessageBox.critical(self, '错误', '请输入规则名并选择对应项', QMessageBox.Yes)
                else:
                    self.rule.append([name, [(i, self.table_header[i]) for i in indexes]])
                    self.refresh_rule_widget()
        else:
            QMessageBox.critical(self, '错误', '请先打开文件', QMessageBox.Yes)

    def remove_rule(self):
        selected_index = self.ui.rules_widget.selectedIndexes()
        if selected_index:
            selected_index = selected_index[0].row()
            self.ui.rule_detail_widget.clear()
            self.rule.pop(selected_index)
            self.refresh_rule_widget()

    def edit_rule(self):
        selected_index = self.ui.rules_widget.selectedIndexes()
        if selected_index:
            selected_index = selected_index[0].row()
            rule = self.rule[selected_index]
            self.add_rule_window = AddRuleMainWindow(self.table_header, rule[0], [rule[1][i][0] for i in range(len(rule[1]))])
            self.add_rule_window.setWindowModality(Qt.ApplicationModal)
            ret = self.add_rule_window.exec_()
            if ret:
                name, indexes = self.add_rule_window.get_result()
                indexes = [item.row() for item in indexes]
                try:
                    self.rule[selected_index][0] = name
                    self.rule[selected_index][1] = [(i, self.table_header[i]) for i in indexes]
                    self.refresh_rule_widget()
                except Exception as e:
                    print(e)


    def open_rule(self):
        if not self.opened_files:
            QMessageBox.critical(self, '错误', '请首先打开文件', QMessageBox.Yes)
        else:
            file_dialog = QFileDialog(self)
            file_dialog.setFileMode(QFileDialog.ExistingFile)
            file_name, _ = file_dialog.getOpenFileName(filter='规则(*.pkl)')
            if file_name:
                ret = self.open_and_check_rule_file(file_name)
                if not ret:
                    QMessageBox.critical(self, '错误', '所选文件不是规则文件或与所打开表格头不一致', QMessageBox.Yes)
                else:
                    self.refresh_rule_widget()

    def open_and_check_rule_file(self, file_path):
        with open(file_path, 'rb') as f:
            data = pickle.load(f)
        if not isinstance(data, list):
            return False
        for item in data:
            if not isinstance(item, list) or len(item) != 2 or not isinstance(item[0], str) or not isinstance(item[1], list):
                return False
            for it in item[1]:
                if not isinstance(it, tuple) or not isinstance(it[0], numbers.Number) or it[1] not in self.table_header:
                    return False
        self.rule = data
        return True

    def save_rule(self):
        if not self.rule:
            QMessageBox.critical(self, '错误', '请首先创建规则', QMessageBox.Yes)
        else:
            file_path, _ = QFileDialog.getSaveFileName(self, '保存规则', './', '规则文件(*.pkl)')
            if file_path:
                with open(file_path, 'wb') as f:
                    pickle.dump(self.rule, f)

    def refresh_rule_widget(self):
        self.ui.rule_detail_widget.clear()
        self.ui.rules_widget.clear()
        for rule in self.rule:
            self.ui.rules_widget.addItem(rule[0])
        if self.rule:
            self.ui.start_process_btn.setEnabled(True)

    def rule_item_clicked(self, item: QListWidgetItem):
        selected_index = self.ui.rules_widget.selectedIndexes()[0].row()
        data = self.rule[selected_index][1]
        self.ui.rule_detail_widget.clear()
        for item in data:
            self.ui.rule_detail_widget.addItem(item[1])

    def start_process(self):
        start_time = self.ui.start_date_edit.date()
        end_time = self.ui.end_date_edit.date()
        if start_time >= end_time:
            QMessageBox.critical(self, '错误', '结束日期要在开始日期之后', QMessageBox.Yes)
        else:
            csv_result = self.generate_csv()
            if csv_result is not None:
                self.generate_score(csv_result, self.ui.modeComboBox.currentIndex())

    def calc_score(self, score_dict):
        try:
            score = 100 * (score_dict['满意'] + score_dict['较满意'] * 0.9 + score_dict['一般'] * 0.7 +
                           score_dict['不满意'] * 0.4) / (
                            score_dict['满意'] + score_dict['较满意'] + score_dict['一般'] + score_dict[
                        '不满意'] + score_dict['很不满意'])
        except ZeroDivisionError as e:
            score = 0
        return score

    def generate_score(self, csv_result, mode):
        if mode == 0:
            total_num = [len(c) * len(self.rule) for c in csv_result]
            total_num = sum(total_num)
            count = 0
            progress = QProgressDialog(self)
            progress.setWindowTitle("正在处理")
            progress.setLabelText("正在生成结果")
            progress.setMinimumDuration(1)
            progress.setWindowModality(Qt.WindowModal)
            progress.setRange(0, total_num)
            dst_path = os.path.dirname(self.opened_files[0])
            for index, file_name in enumerate(self.opened_files):
                data = csv_result[index]
                result = {}
                for item in data:
                    address = item['address']
                    result[address] = {}
                    for key, value in self.rule:
                        score_dict = {'满意': 0, '较满意': 0, '一般': 0, '不满意': 0, '很不满意': 0, '未接触': 0}
                        for _, name in value:
                            try:
                                d = item[name]
                            except KeyError as e:
                                print(file_name, name)
                            d = d.replace("'", "\"")
                            d = json.loads(d)
                            result[address]['样本量'] = sum(d.values())
                            for k, v in d.items():
                                if k in score_dict:
                                    score_dict[k] += v
                        score = self.calc_score(score_dict)
                        result[address][key] = score
                        count += 1
                        progress.setValue(count)
                header = ['address'] + [item[0] for item in self.rule] + ['样本量']
                result_work_book = openpyxl.Workbook()
                result_sheet = result_work_book.active
                for col in range(len(header)):
                    result_sheet.cell(1, col + 1).value = header[col]
                for row_index, (address, v) in enumerate(result.items()):
                    result_sheet.cell(row_index + 2, 1).value = address
                    for col in range(1, len(header)):
                        result_sheet.cell(row_index + 2, col + 1).value = v[header[col]]
                base_file_name = os.path.basename(self.opened_files[index])
                result_work_book.save(os.path.join(dst_path, f"{base_file_name.split('.')[0]}_result.xlsx"))
        elif mode == 1:
            total_num = [len(c) * len(self.rule) for c in csv_result]
            total_num = sum(total_num)
            count = 0
            progress = QProgressDialog(self)
            progress.setWindowTitle('正在处理')
            progress.setLabelText('正在生成结果')
            progress.setMinimumDuration(1)
            progress.setWindowModality(Qt.WindowModal)
            progress.setRange(0, total_num)
            dst_path = os.path.dirname(self.opened_files[0])
            result_workbook = openpyxl.Workbook()
            for index, filename in enumerate(self.opened_files):
                data = csv_result[index]
                for item in data:
                    address = item['address']
                    result_sheet = result_workbook.create_sheet(address, index=index)
                    result_sheet.cell(1, 1).value = address
                    result_sheet.cell(1, 3).value = '满意程度'
                    degree_name = ['满意', '较满意', '一般', '不满意', '很不满意', '未接触']
                    for i in range(len(degree_name)):
                        result_sheet.cell(2, i + 3).value = degree_name[i]
                    result_sheet.cell(1, 9).value = '得分'
                    degree_row_count = [0, ]
                    for rule_index, (key, value) in enumerate(self.rule):
                        result_sheet.cell(3 + sum(degree_row_count), 1).value = key
                        degree_row_count.append(len(value))
                        for value_index, (_, name) in enumerate(value):
                            try:
                                d = item[name]
                            except KeyError as e:
                                print(filename, name)
                            d = d.replace("'", "\"")
                            d = json.loads(d)
                            result_sheet.cell(3 + sum(degree_row_count[:-1]) + value_index, 2).value = name
                            for col in range(len(degree_name)):
                                result_sheet.cell(3 + sum(degree_row_count[:-1]) + value_index, 3 + col).value = d[degree_name[col]]
                                result_sheet.cell(3 + sum(degree_row_count[:-1]) + value_index, 9).value = self.calc_score(d)
                        count += 1
                        progress.setValue(count)
                base_file_name = os.path.basename(self.opened_files[0])
                result_workbook.save(os.path.join(dst_path, f"{base_file_name.split('.')[0]}_result.xlsx"))



    def generate_csv(self):
        dst_path = os.path.dirname(self.opened_files[0])
        sheets = [openpyxl.load_workbook(name)['Sheet1'] for name in self.opened_files]
        total_num = 0
        for sheet in sheets:
            total_num += (sheet.max_row - 1) * (len(self.table_header) - 9)
        progress = QProgressDialog(self)
        progress.setWindowTitle("正在处理")
        progress.setLabelText("正在生成csv")
        progress.setMinimumDuration(1)
        progress.setWindowModality(Qt.WindowModal)
        progress.setRange(0, total_num)
        count = 0
        final_result = []
        for index, sheet in enumerate(sheets):
            num_row = sheet.max_row
            num_column = sheet.max_column
            column_names = ['address']
            for col in range(10, num_column + 1):
                column_names.append(sheet.cell(1, col).value.strip())
            result = {}
            for row in range(2, num_row + 1):
                submit_date = sheet.cell(row, 2).value.strip()
                submit_date = datetime.datetime.strptime(submit_date, self.DATETIME_FORMAT)
                submit_date = QDate(submit_date.year, submit_date.month, submit_date.day)
                start_date = self.ui.start_date_edit.date()
                end_date = self.ui.end_date_edit.date()
                if submit_date < start_date or submit_date > end_date:
                    count += (len(self.table_header) - 9)
                    progress.setValue(count)
                    continue
                address = sheet.cell(row, 9).value.strip()
                if address not in result:
                    result[address] = {}
                for col in range(10, num_column + 1):
                    value = sheet.cell(row, col).value.strip()
                    if col not in result[address]:
                        result[address][col] = {'满意': 0, '较满意': 0, '一般': 0, '不满意': 0, '很不满意': 0,
                                                '未接触': 0}
                    if value in result[address][col].keys():
                        result[address][col][value] += 1
                    count += 1
                    progress.setValue(count)
            if not result:
                QMessageBox.critical(self, '错误', '选择时间范围内没有符合条件数据', QMessageBox.Yes)
                return None
            csv_result = []
            file_name = os.path.basename(self.opened_files[index])
            with open(os.path.join(dst_path, file_name.split('.')[0] + '.csv'), 'w', encoding='utf8') as f:
                writer = csv.DictWriter(f, column_names, lineterminator='\n')
                writer.writeheader()
                for address, data in result.items():
                    write_data = {}
                    write_data['address'] = address
                    for col in range(10, num_column + 1):
                        write_data[column_names[col - 9]] = str(result[address][col])
                    writer.writerow(write_data)
                    csv_result.append(write_data)
            final_result.append(csv_result)
        return final_result

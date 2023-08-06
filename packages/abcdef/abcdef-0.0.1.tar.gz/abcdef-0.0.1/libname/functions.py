# -*- coding: utf-8 -*-
"""
Created on Fri Jun 10 17:01:00 2022

@author: NOKIA T&P Team
"""
'''
import subprocess, sys, os, time 

def install():
    packages = ['paramiko','scp',  'regex', 'ipywidgets', 'IPython', 'tqdm', 'ipyfilechooser', 'datetime']
    for pkg in packages:
        print(f'Installing {pkg}')
        try:
            subprocess.check_call([sys.executable, "-m", "pip", "install", pkg])
        except Exception as _e:
            print(_e)
        #finally:
        #    import importlib
        #    print(f'Importing {pkg}')
        #    importlib.import_module(pkg)
'''

import sys, os, subprocess, pkg_resources, time, socket, threading, math

required = {'paramiko','scp',  'regex', 'ipywidgets', 'ipython', 'tqdm', 'ipyfilechooser', 'openpyxl'}
installed = {pkg.key for pkg in pkg_resources.working_set}
missing = required - installed
try: 
    if missing:
        python = sys.executable
        subprocess.check_call([python, '-m', 'pip', 'install', *missing], stdout=subprocess.DEVNULL)
except Exception as _e:
    print(_e)

import paramiko
from paramiko import SSHClient, AutoAddPolicy
from IPython.display import display, Markdown, clear_output, display_markdown
import regex as re
import ipywidgets as widgets
from scp import SCPClient
from ipyfilechooser import FileChooser
import config
from datetime import datetime
from tqdm.notebook import tqdm, trange
from openpyxl import load_workbook
from ipywidgets import interact, interactive, fixed, interact_manual

#------------------------- Connectivity Part -----------------------
import getpass      #Allows for secure prompting and collection of the user password

def interactive_auth_handler(title, instructions, prompt_list):
    if prompt_list:
        prompt = prompt_list[0][0]
        if "password" in prompt.strip().lower():
           return [getpass.getpass(prompt_list[0][0])]
        elif 'passcode' in prompt.strip().lower():
           code = input(prompt)
           return [code]
        else:
           print(f'Unknown prompt and returning empty list: ----->  {prompt_list[0][0]}')
    else:
        print('Prompt Empty')
        return []


def handler(title, instructions, prompt_list):
        print('entered in handler')
        answers = []
        for prompt_, _ in prompt_list:
            print('prompt is {prompt_}')
            prompt = prompt_.strip().lower()
            if 'password' in prompt:
                password = getpass.getpass(prompt)
                answers.append(password)
            elif 'passcode' in prompt:
                otp = input(prompt)
                answers.append(otp)
            else:
                print('Unknown prompt: {prompt_}')
        return answers


##########################
# SSH Connection helpers #
##########################


def connect_client(host, username, password):
    try:
        client = SSHClient()
        client.set_missing_host_key_policy(AutoAddPolicy())
        client.connect(host, username=username, password=password)
        if client.get_transport() is not None:
            if client.get_transport().is_active():
                #print(f'Connected {host}')
                display_markdown(f'''✅  ***`{host}`*** ➜  Connected</span>''', raw=True)
                
                return client
            else: 
                display_markdown(f''' ✅  ***`{host}`*** ➜  Connected</span>''', raw=True)
                return None
    except paramiko.AuthenticationException:
        print(f'Failed to connect to {host} due to wrong username/password')
        return None
    except Exception as _e:
        display_markdown(f'''➜  ***`{host}`*** Connection Failed ☠</span>''', raw=True)
        print(_e)


def transport(client, dest_addr, local_addr, port, user, password, key=None):
    try:
        vmtransport = client.get_transport()
        dest_addr = (dest_addr, port) #edited#
        local_addr = (local_addr, port) #edited#
        vmchannel = vmtransport.open_channel("direct-tcpip", dest_addr, local_addr)  # changed vmtransport to client
        try:
            jhost = SSHClient()
            jhost.set_missing_host_key_policy(AutoAddPolicy())
            #jhost.load_host_keys('/home/shoaib/.ssh/known_hosts') #disabled#
            if key: jhost.connect(dest_addr, username=user,key_filename=key, sock=vmchannel)
            else: jhost.connect(dest_addr, username=user, password=password, sock=vmchannel)
            
            if jhost.get_transport().is_active():
                #print(f'Transported to {dest_addr} successfully!')
                display_markdown(f'''🔀 Transported  ➜  {dest_addr} ✅ Successfully</span>''', raw=True)
                return jhost
            else:
                print(f'Connection resfused from {dest_addr}!')
                return None
        except Exception as _e:
            print(_e)
    except Exception as _e:
        print(_e)    

def get_credentials():
    try:
        RAS = config.jump_server_IP # input('Please Enter RAS IP!')
        RAS_USER = config.jump_server_USER#input('Please Enter RAS USER!')
        RAS_PASS = config.jump_server_PASS  #input('Please Enter RAS User Pass!')
        if RAS and RAS_USER and RAS_USER:
            RAS_client = connect_client(RAS,RAS_USER, RAS_PASS)
            HIP = config.CMREPO_HIP #input('Please Enter HIP1 IP: ')
            if RAS_client is not None:
                HIP = transport(RAS_client, dest_addr = HIP, local_addr= RAS, port=22, user=config.CMREPO_HIP_USER, password=config.CMREPO_HIP_PASS)
                if HIP is None:
                    print(f'Conection Refused from HIP!')
                REPO1 = config.CMREPO1    #input('Please Enter CMREPO1 IP: ')
                CMREPO1 = transport(RAS_client, dest_addr = REPO1, local_addr= RAS, port=22, user=config.CMREPO1_USER, password=config.CMREPO1_PASS)
                if CMREPO1 is None:
                    print(f'Connection Refused from CMREPO1!')
                REPO2 = config.CMREPO2    #input('Please Enter CMREPO2 IP: ')
                CMREPO2 = transport(RAS_client, dest_addr = REPO2, local_addr= RAS, port=22, user=config.CMREPO2_USER, password=config.CMREPO2_PASS)
                if CMREPO2 is None:
                    print(f'Connection refused from CMREPO2!')

                CBIS = transport(RAS_client, dest_addr = config.CBIS_IP, local_addr= RAS, port=22, user=config.CBIS_USER, password=config.CBIS_PASS)
                if CBIS is None:
                    print(f'Connection refused from CBIS!')
                CBAM = transport(RAS_client, dest_addr = config.CBAM_IP, local_addr= RAS, port=22, user=config.CBAM_USER, password='', key=config.CBAM_KEY)
                if CBAM is None:
                    print(f'Connection refused from CBAM!')
                    
                return HIP,CMREPO1, CMREPO2, CBIS, CBAM
            else: 
                print('Unfortunately Could not establish connection with Jump server!')
                return None, None, None, None, None
        else:
            print('Please provide all credentials for Jump server in config file!')
            return None, None, None, None, None
    except Exception as _e:
        print(_e)


def connect():
    try:
        if 'yes' in config.Jump_server.lower():
            print('Connecting through Jump Server......')
            HIP,CMREPO1, CMREPO2, CBIS, CBAM  = get_credentials()
            if HIP and CMREPO1 and CMREPO2 and CBIS and CBAM:
                return HIP,CMREPO1, CMREPO2, CBIS, CBAM
        else:
            print('Please correct Credentials in config.py file!')
        
        return HIP,CMREPO1, CMREPO2, CBIS, CBAM
    
    except Exception as _e:
        print(_e)


###########
# Helpers #
########### 

def get_ne_version(baseline, target_ne_ver, key=False):
    for k, v in baseline.items():
        if isinstance(v, str) and target_ne_ver.lower() in v.lower():
            return v
    if key:
        ver_list =[key for key, value in baseline.items() if isinstance(value, str) and target_ne_ver.lower() in value.lower()]
    else:
        ver_list =[value for key, value in baseline.items() if isinstance(value, str) and target_ne_ver.lower() in value.lower()]
    return ver_list    


def hex_to_rgb(hex):
    h = hex.lstrip('#')
    return tuple(int(h[i:i+2], 16) for i in (0, 2, 4))

def calculate_distance(cor1, cor2):
    c1 = hex_to_rgb(cor1)
    c2 = hex_to_rgb(cor2)
    return math.sqrt((c1[0]-c2[0])**2+(c1[1]-c2[1])**2+(c1[2]-c2[2])**2)



def color_test(sh, target_ver, target_ne, target):
    """
    gets row and col from .xlsx sheet, parse cell location and checks color
    """
    #print(f'target version {target_ver}\n target ne {target_ne}')
    col = re.match("([a-zA-Z]+)([0-9]+)", target_ne).group(1)
    row = re.match("([a-zA-Z]+)([0-9]+)", target_ver).group(2)
    color_in_hex = sh[col+row].fill.start_color.index
    
    dist = calculate_distance(color_in_hex, '#00FF00')
 
    if dist>255:
        display_markdown(f'''<span style='background :lightgreen' > Compatible `with` ***`{target.get(target_ne)}`*** </span> 😊 ''', raw=True)
    else:
        display_markdown(f'''<span style='background :orange' > Not Compatible `with` ***`{target.get(target_ne)}`*** </span> 😔 ''', raw=True)

def get_ne_version(baseline, target_ne_ver, key=False):
    for k, v in baseline.items():
        if isinstance(v, str) and target_ne_ver.lower() in v.lower():
            return v
    if key:
        ver_list =[key for key, value in baseline.items() if isinstance(value, str) and target_ne_ver.lower() in value.lower()]
    else:
        ver_list =[value for key, value in baseline.items() if isinstance(value, str) and target_ne_ver.lower() in value.lower()]
    return ver_list    


########################
# get current versions #
########################


def get_current_vnf_versions(HIP,CBIS, CBAM):
    """
    Login into each network element and returns version string
    """
    # CMREPO Version
    ver_CMREPO = ''
    stdin, stdout, stderr = HIP.exec_command('cmcli info')
    output = stdout.readlines()
    cmrepo_flag = False
    if output:
        for line in output:
            if 'type:' in line.lower() and 'cmrepo' in line.lower():
                cmrepo_flag = True
            if cmrepo_flag and 'Active Release:' in line:

                if 'active release:' in line.lower():
                    try:
                        ver_CMREPO = re.search("(\d+\.?\d*)", line).group(1)
                        #print(f'CMREPO Current active release is {ver_CMREPO}')
                    except AttributeError:
                        ver_CMREPO = re.search("(\d+\.?\d*)", line) 
                    #return re.match(r'\d+\.?\d*', line).group(1)
                    break;

    else: 
        print(f'We can not get version for CMREPO!\nError: \n{stderr}')
    
    # CBIS version
    ver_CBIS=''
    stdin, stdout,stderr = CBIS.exec_command('openstack cbis version')
    output = stdout.readlines()
    if output:
        for line in output:
            if 'build' in line.lower():
                ver_CBIS = re.search('(\d+\.?\d+)', line).group(1)
                temp_ver = re.search('(\d+\.?\d+)', ver_CBIS).group(1)
                ver = re.search('(\.?\d+)', temp_ver).group(1)
                if ver: ver_CBIS = 'CBIS '+ver+'A SW' 
                else: ver_CBIS = 'CBIS '+ver_CBIS+'A SW'
                #print(f'CBIS Current version is {ver_CBIS}')
    else:
        error = stderr.read().decode('utf8')
        print(f'\nCBIS Error:\n {error}')
    #CBAM version
    ver_CBAM = ''
    stdin, stdout,stderr = CBAM.exec_command( 'cbam --version')
    output = stdout.readline()
    
    if output:
        try:
            ver_CBAM = re.search("(\d+\.\d+)", output).group(1) 
            #print(f'CBAM Current version is {ver_CBAM}')
        except AttributeError:
            ver_CBAM = re.search("(\d+\.\d+\.\d+\.\d+)", output).group(1)
    else:
        print(f'\nCBAM Error:\n{stderr.read().decode()}')
    
    
    if ver_CBAM and ver_CBIS and ver_CMREPO:
        return 'CBAM '+ver_CBAM +' SW', ver_CBIS, ver_CMREPO


########################
# Compatibility Matrix #
########################


def comp_matrix(target_ne_ver,CMREPO_HIP,CBIS,CBAM,ne_versions=None, sheet='Compa Matrix', path='Cloud_matrix.xlsx'):
    """
    reads .xlsx file and check compatibility
    """
    #temp for showing demo
    ver_CBAM, ver_CBIS, ver_CMREPO = get_current_vnf_versions(CMREPO_HIP,CBIS,CBAM)
    ne_versions = [ver_CBAM, ver_CBIS]
    
    if isinstance(path, str) and path.endswith('.xlsx'):
        excel_file = path
        wb = load_workbook(excel_file, data_only = True)
        sh = wb[sheet]

        baseline, target = {}, {}
        for key, *values in sh.iter_rows():
            baseline[key.coordinate] = key.value
            if key.value == ' ': #key.key.coordinate == 'A2': # key.value is None:
                for target_version in values:
                    target[target_version.coordinate] = target_version.value
        
        base  =[value for key, value in baseline.items() if isinstance(value, str) and target_ne_ver.lower() in value.lower()]
        available_target = widgets.Dropdown(options = base) # returns dropdown of listed versions
        _netAct = [value for key, value in target.items() if isinstance(value, str) and 'netact' in value.lower()]
        _netAct_versions = widgets.Dropdown(options = _netAct)
        
        @interact(CFX_Target = available_target, NetAct = _netAct_versions) # change1
        def print_version(CFX_Target, NetAct):
            Target = available_target
            target_ver = ''.join(str(x) for x in [key for key, value in baseline.items() if isinstance(value, str) and Target.value.lower()==value.lower()])
            
            NetAct = ''.join(str(x) for x in [key for key, value in target.items() if isinstance(value, str) and NetAct.lower()==value.lower()])
            color_test(sh, target_ver, NetAct.strip(), target)
            
            for ne in ne_versions:
                target_ne = [key for key, value in target.items() if isinstance(value, str) and ne.lower().strip() == value.lower()]
                target_ne = ''.join(str(x) for x in target_ne)
                color_test(sh, target_ver, target_ne, target)
    
    else: display_markdown(f'''  Please select ***`.xlsx`*** file 😔''', raw=True)

#-------------------------  Command Execution  ----------------------

def exec_command(client, cmd, vm_type=None, vm_ip=None):
    try:
        
        stdin, stdout, stderr = client.exec_command(cmd)
        out = stdout.read().decode('utf-8')
        error = stderr.read().decode('utf8')
        
        if not os.path.exists('logs_trail.txt'):
                with open("logs_trail.txt", "x") as file:
                    file.write('########################################\n##------------- LOGS TRAIL -----------##\n########################################\n')
        with open("logs_trail.txt", "a", encoding="utf-8") as text_file:   
            if out:
                if vm_type and vm_ip:
                    display_markdown(f'''<span style='background :lightgreen' > 🔀 Execution happening on ⇨ ***`{vm_type}`*** type VM against IP ⇨ ***`{vm_ip}`***  ➜  OUTPUT : ✅🔽  </span> \n\n ``` {out} ```''', raw=True)
                else: 
                    print(f'OUTPUT:\n {out}\n')
                text_file.write(f"\n####################\n#***** OUTPUT *****#\n####################\nStep Started at: {datetime.now()} \nCommands------->>>>>>>>> '{cmd} <<<<<<<<<  \n\n{out} \nStep ended at: {datetime.now()}")
            if error:
                if vm_type and vm_ip:
                    display_markdown(f'''<span style='background :red' >🔀 Execution happening on ⇨ ***`{vm_type}`*** type VM against IP ⇨ ***`{vm_ip}`***    ➜  ERROR : 📛🔽  </span> \n\n ``` {error} ```''', raw=True)
                else: 
                    print(f'ERROR:\n {error}\n')
                text_file.write(f"\n\n###################\n#***** ERROR *****#\n###################\n\nError Occured at: {datetime.now()} time\n{error}\n")   
    except Exception as _e:
        print(_e)
#-------------------------  helping Function  -----------------------

def progress4(filename, size, sent, peername):
    sys.stdout.write("(%s:%s) %s's progress: %.2f%%   \r" % (peername[0], peername[1], filename, float(sent)/float(size)*100) )

def select():
    folder = FileChooser(os.getcwd())
    display(folder)
    return folder

def get(user, file_name, folder):
    try:
        if folder.selected:
            scp = SCPClient(user.get_transport(), progress4=progress4)
            scp.get(file_name, folder.selected)
        else:
            print('Please select file path!')
    except Exception as _e:
        print(_e)

def cmrepo_prep(user, prep):
    try:
        if prep.selected:
            print('Processing Prepration, Please wait.....!')
            cmd_list = 'cmcli info; cmcli backup; crontab -l; ls -ltr /etc; ls -ltr /opt; ls -ltr /home; ls -ltr /tmp/li; ls -ltr /home/iserver/linux/http/static/swcli.py; ls -ls /tmp/LI/*.*; netstat -ne; netstat -nr'
            # temporary fo demo
            _,out1, _ = user.exec_command(cmd_list)
            out1 = out1.read().decode('utf8')
            with open("prep.txt", "w") as file:
                file.write(out1)
            exec_command(user, cmd_list)
            get(user, '/opt/kms/vmon/DND/nb/kms.tar.gz', prep)
        else:
            print('Please select folder to store kms.tar.gz file!')
    except Exception as e:
        print(e)

def progress_bar(start, end):
    for i in tqdm(range(start, end)):
        time.sleep(0.00001)


def select_backup_path():
    local_file = select(); local_file.title=' Select CmrepoDataMigration.tar from Local drive'
    backup_path = select(); backup_path.title=' Folder to store CMREPO2 backup file'
    return local_file, backup_path

def compare(client, s1,s2):
    _,out1, _ = client.exec_command(s1)
    _, out2, _ = client.exec_command(s2)
    out1 =out1.readlines()
    out2 = out2.readlines()
    for line in out1:
        #print(line)
        if line not in out2:
            display_markdown(f'''<span style='color :blue' > {line} </span> ''', raw=True)
    for line in out2:
        if line not in out1:
            display_markdown(f'''<span style='color :red' > {line} </span> ''', raw=True)
def compare_prep(user):
    cmd_list = 'cmcli info; cmcli backup; crontab -l; ls -ltr /etc; ls -ltr /opt; ls -ltr /home; ls -ltr /tmp/li; ls -ltr /home/iserver/linux/http/static/swcli.py; ls -ls /tmp/LI/*.*; netstat -ne; netstat -nr'
    _,out1, _ = user.exec_command(cmd_list)
    out1 = out1.readlines()
    with open("prep.txt") as file:
        out2 = file.readlines()
    for line in out1:
        #print(line)
        if line not in out2:
            display_markdown(f'''<span style='color :blue' > {line} </span> ''', raw=True)
    for line in out2:
        if line not in out1:
            display_markdown(f'''<span style='color :red' > {line} </span> ''', raw=True)

def parallel_executions(CFX, cmd, vm_type, vm_ip):
    ip = transport(CFX, dest_addr = vm_ip, local_addr= '192.168.130.50', port=22, user='root', password='yt_xk39b')
    exec_command(ip,cmd,vm_type, vm_ip)

#-------------------------  Pre-Checks  ----------------------------

def write_into_file(step,check_no, out, error, cmd):
    if not os.path.exists('logs_trail.txt'):
        with open("logs_trail.txt", "x") as file:
            file.write('########################################\n##------------- LOGS TRAIL -----------##\n########################################\n')
    with open("logs_trail.txt", "a", encoding="utf-8") as text_file:
        if out:
            text_file.write(f"\n###########################\n#***** {step} ({check_no}) *****#\n###########################\nCommands Step-> **{cmd}** \n\n{out}\n")
        if error:
            text_file.write(f"\n\n###################\n#***** ERROR *****#\n###################\n\n{error}\n")     

def pre_check(client):
    if client:
        try:
            #--------------------------Check-1----------------------------#
            cmd = 'systemctl status cmserver'
            stdin, stdout, stderr = client.exec_command(cmd)
            out = stdout.read().decode('utf8')
            error = stderr.read().decode('utf8')
            write_into_file('PRE-CHECK',1, out, error, cmd)
            
            print(f'\n \n Check No.1 : systemctl status cmserver \n{out}' )
            if 'active (running) since' in out and not error:
                #service = re.search(r'([/][^;]*)', out).group(1)
                active = re.search(r'(active[^;]*)',out).group(1)
                display_markdown('''<span style='background :lightgreen' > Check Passed! </span> ''', raw=True)
                print(f'cmserver is {active}')
                progress_bar(0, 100)
            else:
                display_markdown('''<span style='background :red' > Check Failed! </span> ''', raw=True)
                print('Please check cmserver is not active! -----> Check Failed!')
                print(error)
            
            #--------------------------Check-2----------------------------#
            cmd = 'systemctl status swagent'
            stdin, stdout, stderr = client.exec_command(cmd)
            out = stdout.read().decode('utf8')
            error = stderr.read().decode('utf8')
            write_into_file('PRE-CHECK',2, out, error, cmd)
            
            print(f'\n \n Check No.2 : systemctl status swagent \n{out}' )
            if 'active (running) since' in out and not error:
                #service = re.search(r'([/][^;]*)', out).group(1)
                active = re.search(r'(active[^;]*)',out).group(1)
                display_markdown('''<span style='background :lightgreen' > Check Passed! </span> ''', raw=True)
                print(f'swagent is {active} ---> Check Passed ')
                progress_bar(0, 100)
            else:
                display_markdown('''<span style='background :red' > Check Failed! </span> ''', raw=True)
                print(f'Please check swagent is not active! ---> Check Failed!')
                print(error)
            
            #--------------------------Check-3----------------------------#
            cmd = 'ps -ef |grep -e cmserver -e cmrelay'
            stdin, stdout, stderr = client.exec_command(cmd)
            out = stdout.read().decode('utf8')
            error = stderr.read().decode('utf8')
            write_into_file('PRE-CHECK',3, out, error, cmd)
            
            print(f'\n \n Check No.3 : ps -ef |grep -e cmserver -e cmrelay \n{out}' )
            # cmserver->1, cmrelay ->2, bash
            if  re.search('cmserver |cmrelay|grep -e cmserver -e cmrelay',out) and not error:
                display_markdown('''<span style='background :lightgreen' > Check Passed! </span> ''', raw=True)
                print(f'Both cmrelay and cmserver appeared  ---> Check Passed ')
                progress_bar(0, 100)
            else:
                display_markdown('''<span style='background :red' > Check Failed! </span> ''', raw=True)
                print(f'Please check cmrelay and cmserver did not appeared! ---> Check Failed')
                print(error)
            
            #--------------------------Check-4----------------------------#
            cmd = '/opt/cmbase/server/bin/GetRepomode.sh'
            stdin, stdout, stderr = client.exec_command(cmd)
            out = stdout.read().decode('utf-8')
            error = stderr.read().decode('utf-8')
            write_into_file('PRE-CHECK',4, out, error, cmd)
            
            print(f'\n \n Check No. 4 : /opt/cmbase/server/bin/GetRepomode.sh \n{out}' )
            if  re.search('Primary| Secondary| master| slave',out):
                display_markdown('''<span style='background :lightgreen' > Check Passed! </span> ''', raw=True)
                print(f'Both Modes (Primary and Secondary) are running  ---> Check Passed ')
                progress_bar(0, 100)
            else: 
                display_markdown('''<span style='background :red' > Check Failed! </span> ''', raw=True)
                print(f'Please check Modes are not running! ---> Check Failed')
                print(error)
            
            #--------------------------Check-5----------------------------#
            out = ''
            cmd = '/opt/SMAW/bin/execRTPenv status1'
            stdin, stdout, stderr = client.exec_command(cmd)
            stdout = stdout.readlines()
            for line in stdout:
                out= out + line
            error = stderr.read().decode('utf8')
            write_into_file('PRE-CHECK', 5, out, error, cmd)
            
            print(f'\n \n Check No. 5 : /opt/SMAW/bin/execRTPenv status1 \n{out}' )
            flag=False
            e_stdin, e_stdout, e_stderr = client.exec_command('/opt/SMAW/bin/execRTPenv status1 -e')
            
            for line in stdout:
                if 'run' not in line and e_stdout=='':
                    flag=True
                     
            if not flag and not error:
                display_markdown('''<span style='background :lightgreen' > Check Passed! </span> ''', raw=True)
                print(f'All processes are running  ---> Check Passed ')
                progress_bar(0, 100)
            else:
                display_markdown('''<span style='background :red' > Check Failed! </span> ''', raw=True)
                print(f'Please check some process are not running! ---> Check Failed')
                print(e_stderr.read().decode('utf8'))
                print(error)
            
            #--------------------------Check-6--------------------------#
            out = ''
            cmd = 'netstat -an |grep -e 7070 -e 7043'
            stdin, stdout, stderr = client.exec_command(cmd)
            stdout = stdout.readlines()
            for line in stdout:
                out= out + line
            error = stderr.read().decode('utf8')
            write_into_file('PRE-CHECK', 6, out, error, cmd)
            
            print(f'\n \n Check No. 6 : netstat -an |grep -e 7070 -e 7043 \n{out}')
            
            flag = False
            for line in stdout:
                if all(word in line for word in ['7070', 'LISTEN']):
                    display_markdown('''<span style='background :lightgreen' > Check Passed! </span> ''', raw=True)
                    print(f'7070 ports are Listining  ---> Check Passed ')
                    progress_bar(0, 100); flag=True
                    break
            if not flag:
                display_markdown('''<span style='background :red' > Check Failed! </span> ''', raw=True)
                print(f'Please check some ports are not listning at 7070 and 7043! ---> Check Failed')
                print(error)

            
            #--------------------------Check-7--------------------------#
            out = ''
            cmd = 'netstat -an |grep -e 8080 -e 8443'
            stdin, stdout, stderr = client.exec_command(cmd)
            stdout = stdout.readlines()
            for line in stdout:
                out= out + line
            error = stderr.read().decode('utf8')
            write_into_file('PRE-CHECK', 7, out, error, cmd)
            
            print(f'\n \n Check No. 7 : netstat -an |grep -e 8080 -e 8443 \n{out}' )
            
            flag =False
            for line in stdout:
                if all(word in line for word in ['8080', 'LISTEN']) or all(word in line for word in ['8443', 'LISTEN']):
                    display_markdown('''<span style='background :lightgreen' > Check Passed! </span> ''', raw=True)
                    print(f'All ports are Listining at 8080 and 8443 ---> Check Passed ')
                    progress_bar(0, 100); flag=True

            if not flag:
                display_markdown('''<span style='background :red' > Check Failed! </span> ''', raw=True)
                print(f'Please check some ports are not listning! ---> Check Failed')
                print(error)
            
            #--------------------------Check-8--------------------------#
            out = ''
            cmd = 'netstat -an |grep 9997'
            stdin, stdout, stderr = client.exec_command(cmd)
            stdout = stdout.readlines()
            for line in stdout:
                out= out + line
            error = stderr.read().decode('utf8')
            write_into_file('PRE-CHECK', 8, out, error, cmd)
            
            print(f'\n \n Check No.8 : netstat -an |grep 9997 \n{out}' )
            
            flag =False
            for line in stdout:
                if all(word in line for word in ['9997', 'LISTEN']):
                    display_markdown('''<span style='background :lightgreen' > Check Passed! </span> ''', raw=True)
                    print(f'All ports are Listining at 9997 ---> Check Passed ')
                    progress_bar(0, 100); flag=True
            if not flag:
                display_markdown('''<span style='background :red' > Check Failed! </span> ''', raw=True)
                print(f'Please check 9997 ports is not listning! ---> Check Failed')
                print(error)

            #--------------------------Check-9--------------------------#
            cmd = 'sudo python /opt/cmbase/server/scripts/CmRepoHealth_Check.py'
            stdin, stdout, stderr = client.exec_command(cmd)
            stdout = stdout.readlines()
            
            flag=False
            out = ''
            failed_check = ''
            for line in stdout:
                if 'FAILED' in line:
                    flag=True
                    failed_check += line
                out=out+line
            error = stderr.read().decode('utf8')
            write_into_file('PRE-CHECK', 9, out, error, cmd)
            
            print(f'\n \n Check No.9 : CmRepoHealth_Check.py \n{out}' )
            
                
            if flag:
                display_markdown('''<span style='background :red' > Check Failed! </span> ''', raw=True)
                print(f'--------> Please troubleshoot this check and continue.........! \n {failed_check}')
                print(error)        
            else:
                display_markdown('''<span style='background :lightgreen' > Check Passed! </span> ''', raw=True)
                print(f'Health is Good and Log file exists ---> Check Passed ')
                progress_bar(0, 100)
            
            #--------------------------Check-10--------------------------#
            #cmd = 'cat /etc/passwd; cat /etc/group'
            #stdin, stdout, stderr = client.exec_command(cmd)
            #out = stdout.read().decode('utf8')
            #error = stderr.read().decode('utf8')
            #write_into_file('PRE-CHECK', 10, out, error, cmd)
            
            #print(f'\n \n Check No. 10 : cat /etc/passwd; cat /etc/group \n{out}')
            #display_markdown('''<span style='background :lightgreen' > Check Passed! </span> ''', raw=True)
           
            
        except Exception as _e:
            print(_e)
    else:
        print('Only rtp99 user is allowed to execute pre-check tests')

#-------------------post checks same as pre-checks-------------

def post_check(client):
    if client:
        try:
            #--------------------------Check-1----------------------------#
            cmd = 'systemctl status cmserver'
            stdin, stdout, stderr = client.exec_command(cmd)
            out = stdout.read().decode('utf8')
            error = stderr.read().decode('utf8')
            write_into_file('PRE-CHECK',1, out, error, cmd)
            
            print(f'\n \n Check No.1 : systemctl status cmserver \n{out}' )
            if 'active (running) since' in out and not error:
                #service = re.search(r'([/][^;]*)', out).group(1)
                active = re.search(r'(active[^;]*)',out).group(1)
                display_markdown('''<span style='background :lightgreen' > Check Passed! </span> ''', raw=True)
                print(f'cmserver is {active}')
                progress_bar(0, 100)
            else:
                display_markdown('''<span style='background :red' > Check Failed! </span> ''', raw=True)
                print('Please check cmserver is not active! -----> Check Failed!')
                print(error)
            
            #--------------------------Check-2----------------------------#
            cmd = 'systemctl status swagent'
            stdin, stdout, stderr = client.exec_command(cmd)
            out = stdout.read().decode('utf8')
            error = stderr.read().decode('utf8')
            write_into_file('PRE-CHECK',2, out, error, cmd)
            
            print(f'\n \n Check No.2 : systemctl status swagent \n{out}' )
            if 'active (running) since' in out and not error:
                #service = re.search(r'([/][^;]*)', out).group(1)
                active = re.search(r'(active[^;]*)',out).group(1)
                display_markdown('''<span style='background :lightgreen' > Check Passed! </span> ''', raw=True)
                print(f'swagent is {active} ---> Check Passed ')
                progress_bar(0, 100)
            else:
                display_markdown('''<span style='background :red' > Check Failed! </span> ''', raw=True)
                print(f'Please check swagent is not active! ---> Check Failed!')
                print(error)
            
            #--------------------------Check-3----------------------------#
            cmd = 'ps -ef |grep -e cmserver -e cmrelay'
            stdin, stdout, stderr = client.exec_command(cmd)
            out = stdout.read().decode('utf8')
            error = stderr.read().decode('utf8')
            write_into_file('PRE-CHECK',3, out, error, cmd)
            
            print(f'\n \n Check No.3 : ps -ef |grep -e cmserver -e cmrelay \n{out}' )
            # cmserver->1, cmrelay ->2, bash
            if  re.search('cmserver |cmrelay|grep -e cmserver -e cmrelay',out) and not error:
                display_markdown('''<span style='background :lightgreen' > Check Passed! </span> ''', raw=True)
                print(f'Both cmrelay and cmserver appeared  ---> Check Passed ')
                progress_bar(0, 100)
            else:
                display_markdown('''<span style='background :red' > Check Failed! </span> ''', raw=True)
                print(f'Please check cmrelay and cmserver did not appeared! ---> Check Failed')
                print(error)
            
            #--------------------------Check-4----------------------------#
            cmd = '/opt/cmbase/server/bin/GetRepomode.sh'
            stdin, stdout, stderr = client.exec_command(cmd)
            out = stdout.read().decode('utf-8')
            error = stderr.read().decode('utf-8')
            write_into_file('PRE-CHECK',4, out, error, cmd)
            
            print(f'\n \n Check No. 4 : /opt/cmbase/server/bin/GetRepomode.sh \n{out}' )
            if  re.search('Primary| Secondary| master| slave',out):
                display_markdown('''<span style='background :lightgreen' > Check Passed! </span> ''', raw=True)
                print(f'Both Modes (Primary and Secondary) are running  ---> Check Passed ')
                progress_bar(0, 100)
            else: 
                display_markdown('''<span style='background :red' > Check Failed! </span> ''', raw=True)
                print(f'Please check Modes are not running! ---> Check Failed')
                print(error)
            
            #--------------------------Check-5----------------------------#
            out = ''
            cmd = '/opt/SMAW/bin/execRTPenv status1'
            stdin, stdout, stderr = client.exec_command(cmd)
            stdout = stdout.readlines()
            for line in stdout:
                out= out + line
            error = stderr.read().decode('utf8')
            write_into_file('PRE-CHECK', 5, out, error, cmd)
            
            print(f'\n \n Check No. 5 : /opt/SMAW/bin/execRTPenv status1 \n{out}' )
            flag=False
            e_stdin, e_stdout, e_stderr = client.exec_command('/opt/SMAW/bin/execRTPenv status1 -e')
            
            for line in stdout:
                if 'run' not in line and e_stdout=='':
                    flag=True
                     
            if not flag and not error:
                display_markdown('''<span style='background :lightgreen' > Check Passed! </span> ''', raw=True)
                print(f'All processes are running  ---> Check Passed ')
                progress_bar(0, 100)
            else:
                display_markdown('''<span style='background :red' > Check Failed! </span> ''', raw=True)
                print(f'Please check some process are not running! ---> Check Failed')
                print(e_stderr.read().decode('utf8'))
                print(error)
            
            #--------------------------Check-6--------------------------#
            out = ''
            cmd = 'netstat -an |grep -e 7070 -e 7043'
            stdin, stdout, stderr = client.exec_command(cmd)
            stdout = stdout.readlines()
            for line in stdout:
                out= out + line
            error = stderr.read().decode('utf8')
            write_into_file('PRE-CHECK', 6, out, error, cmd)
            
            print(f'\n \n Check No. 6 : netstat -an |grep -e 7070 -e 7043 \n{out}')
            
            flag = False
            for line in stdout:
                if all(word in line for word in ['7070', 'LISTEN']):
                    display_markdown('''<span style='background :lightgreen' > Check Passed! </span> ''', raw=True)
                    print(f'7070 ports are Listining  ---> Check Passed ')
                    progress_bar(0, 100); flag=True
                    break
            if not flag:
                display_markdown('''<span style='background :red' > Check Failed! </span> ''', raw=True)
                print(f'Please check some ports are not listning at 7070 and 7043! ---> Check Failed')
                print(error)

            
            #--------------------------Check-7--------------------------#
            out = ''
            cmd = 'netstat -an |grep -e 8080 -e 8443'
            stdin, stdout, stderr = client.exec_command(cmd)
            stdout = stdout.readlines()
            for line in stdout:
                out= out + line
            error = stderr.read().decode('utf8')
            write_into_file('PRE-CHECK', 7, out, error, cmd)
            
            print(f'\n \n Check No. 7 : netstat -an |grep -e 8080 -e 8443 \n{out}' )
            
            flag =False
            for line in stdout:
                if all(word in line for word in ['8080', 'LISTEN']) or all(word in line for word in ['8443', 'LISTEN']):
                    display_markdown('''<span style='background :lightgreen' > Check Passed! </span> ''', raw=True)
                    print(f'All ports are Listining at 8080 and 8443 ---> Check Passed ')
                    progress_bar(0, 100); flag=True

            if not flag:
                display_markdown('''<span style='background :red' > Check Failed! </span> ''', raw=True)
                print(f'Please check some ports are not listning! ---> Check Failed')
                print(error)
            
            #--------------------------Check-8--------------------------#
            out = ''
            cmd = 'netstat -an |grep 9997'
            stdin, stdout, stderr = client.exec_command(cmd)
            stdout = stdout.readlines()
            for line in stdout:
                out= out + line
            error = stderr.read().decode('utf8')
            write_into_file('PRE-CHECK', 8, out, error, cmd)
            
            print(f'\n \n Check No.8 : netstat -an |grep 9997 \n{out}' )
            
            flag =False
            for line in stdout:
                if all(word in line for word in ['9997', 'LISTEN']):
                    display_markdown('''<span style='background :lightgreen' > Check Passed! </span> ''', raw=True)
                    print(f'All ports are Listining at 9997 ---> Check Passed ')
                    progress_bar(0, 100); flag=True
            if not flag:
                display_markdown('''<span style='background :red' > Check Failed! </span> ''', raw=True)
                print(f'Please check 9997 ports is not listning! ---> Check Failed')
                print(error)

            #--------------------------Check-9--------------------------#
            cmd = 'sudo python /opt/cmbase/server/scripts/CmRepoHealth_Check.py'
            stdin, stdout, stderr = client.exec_command(cmd)
            stdout = stdout.readlines()
            
            flag=False
            out = ''
            failed_check = ''
            for line in stdout:
                if 'FAILED' in line:
                    flag=True
                    failed_check += line
                out=out+line
            error = stderr.read().decode('utf8')
            write_into_file('PRE-CHECK', 9, out, error, cmd)
            
            print(f'\n \n Check No.9 : CmRepoHealth_Check.py \n{out}' )
            
                
            if flag:
                display_markdown('''<span style='background :red' > Check Failed! </span> ''', raw=True)
                print(f'--------> Please troubleshoot this check and continue.........! \n {failed_check}')
                print(error)        
            else:
                display_markdown('''<span style='background :lightgreen' > Check Passed! </span> ''', raw=True)
                print(f'Health is Good and Log file exists ---> Check Passed ')
                progress_bar(0, 100)
            
            #--------------------------Check-10--------------------------#
            #cmd = 'cat /etc/passwd; cat /etc/group'
            #stdin, stdout, stderr = client.exec_command(cmd)
            #out = stdout.read().decode('utf8')
            #error = stderr.read().decode('utf8')
            #write_into_file('PRE-CHECK', 10, out, error, cmd)
            
            #print(f'\n \n Check No. 10 : cat /etc/passwd; cat /etc/group \n{out}')
            #display_markdown('''<span style='background :lightgreen' > Check Passed! </span> ''', raw=True)
           
            
        except Exception as _e:
            print(_e)
    else:
        print('Only rtp99 user is allowed to execute pre-check tests')





#-------------------------  UpGrade  ------------------------------

def put(user, local, remote):
    try:
        if local.selected:
            scp = SCPClient(user.get_transport(), progress4=progress4)
            scp.put(local.selected, remote)
        else:
            print('Please select path relevent tar file!')
    except Exception as _e:
        print(_e)

def get_backup(user, local):
    try:
        if local.selected:
            with SCPClient(user.get_transport(),progress4=progress4, sanitize=lambda x: x) as scp:
                        scp.get('/Tspswrepo/CMREPO/backup*.tar.gz', local.selected)
        else:
            print('Please select relevent backup file which ends with tar.gz')
    except Exception as _e:
        print(_e)

def auto_bakup_file(user, local_file, backup_path): 
    try:
        if local_file.selected and backup_path.selected:
            remote = r'/Tspswrepo/'
            scp = SCPClient(user.get_transport(), progress4=progress4)
            scp.put(local_file.selected, remote)
            exec_command(user, 'cd /Tspswrepo; tar -xf /Tspswrepo/CmrepoDataMigration.tar; sh /Tspswrepo/SwRepoDataMigration.sh --backup; ls -l /Tspswrepo/CMREPO')
            get_backup(user, backup_path)
        else:
            print('Please select CmrepoDataMigration.tar and backup folder path')
    except Exception as _e:
        print(_e)

def get_repo_info(user):
    cmd = '/opt/cmbase/server/bin/GetRepomode.sh'
    stdin, stdout, stderr = user.exec_command(cmd)
    out = stdout.readlines()
    for line in out:
        if 'Modes on' in line:
            machine = re.search(r'Modes on (.*):', line).group(1)
        if 'Server :' in line:
            if 'Primary' in line:
                primary = machine
                machine = ''
            elif 'Secondary' in line:
                secondary = machine
                machine = ''
        #print(line)
    #print(f'Primary Repo: {primary}\nSecondary Repo: {secondary}')
    return primary

def update_version_string(user, version):
    # version update
    stdin, stdout, stderr = user.exec_command('grep -i "UniqueId" /tspinst/scripts/aiParameter.sh -w')
    out = stdout.read().decode('utf-8')
    out1 = re.search(r'UniqueId=(.*)', out).group(1) #gets UniqueID
    out2 = re.search(r'([\d.]+)', out).group(1) #gets version no.
    val = out1.replace(out2, version)
    primary = get_repo_info(user)
    
    val =  val.replace(re.search(r'[\d.]+\/([A-Za-z0-9_]+)', val).group(1), primary) 
    #print('\n', out1, '\n Version = ',out2,'\n', val)
    return val

def update_version(user, version):
    file_path = '/opt/cmbase/common/config/config/input-path.conf'
    sftp = paramiko.SSHClient.open_sftp(user)
    with sftp.open(file_path, "r") as file:
        data= file.readlines()
    
    updated_version_str = update_version_string(user, version)
    updated_version_str = 'CMREPO_UUID="'+ updated_version_str+'"'
    print(updated_version_str)
    
    updated = []
    for line in data:
        if 'CMREPO_UUID=' in line:
            if not line.startswith('#'):
                #line= re.sub(r'CMREPO_UUID=[.*]+', updated_version_str, line)
                line = updated_version_str
        updated.append(line)
    
    for i in updated:
        print(i)
    with sftp.open(file_path, 'w') as file:
        file.writelines(updated)


#-------------------------  CFX Connectivity  -----------------------

def get_CFX_credentials():
    try: 
        CFX = config.CFX_IP   #input('Please Enter CFX-OAM IP: ')
        CFX_PASS =config.CFX_PASS    #'yt_xk39b' # ask for default store
        CFX_OAM = connect_client(CFX,config.CFX_USER , CFX_PASS)
        #REPO1 = input('Please Enter CMREPO1 IP: ')
        #CMREPO1 = connect_client(REPO1,'root' , HIP_PASS)
        #CMREPO1 = transport(RAS_client, dest_addr = REPO1, local_addr= RAS, port=22, user='root', password=RAS_PASS)
        #REPO2 = input('Please Enter CMREPO2 IP: ')
        #CMREPO2 = connect_client(REPO2,'root' , HIP_PASS)
        #CMREPO2 = transport(RAS_client, dest_addr = REPO2, local_addr= RAS, port=22, user='root', password=RAS_PASS)
        return CFX_OAM
    except Exception as _e:
        print(_e)

def get_CFX_RAS_credentials():
    RAS =config.jump_server_IP    #input('Please Enter RAS IP!')
    RAS_USER =config.jump_server_USER  #input('Please Enter RAS USER!')
    RAS_PASS = config.jump_server_PASS   #input('Please Enter RAS User Pass!')
    if RAS and RAS_USER and RAS_USER:
        RAS_client = connect_client(RAS,RAS_USER, RAS_PASS)
        CFX_OAM = config.CFX_IP   #input('Please Enter CFX-OAM IP: ')
        CFX_OAM = transport(RAS_client, dest_addr = CFX_OAM, local_addr= RAS, port=22, user='root', password=RAS_PASS)
        #REPO1 = input('Please Enter CMREPO1 IP: ')
        #CMREPO1 = transport(RAS_client, dest_addr = REPO1, local_addr= RAS, port=22, user='root', password=RAS_PASS)
        #REPO2 = input('Please Enter CMREPO2 IP: ')
        #CMREPO2 = transport(RAS_client, dest_addr = REPO2, local_addr= RAS, port=22, user='root', password=RAS_PASS)
        return CFX_OAM
    else:
        print('Please provide all credentials for Jump server!')

def connect_CFX():
    try:
        if 'yes' in config.Jump_server.lower():
            print('Connecting through Jump Server')
            CFX_OAM = get_CFX_RAS_credentials()
            return CFX_OAM
        elif 'no' in config.Jump_server.lower():
            print('Connecting without Jump Server')
            CFX_OAM = get_CFX_credentials()
            return CFX_OAM
        else:
            print('Please check Credentials!')
    
    except Exception as _e:
        print(_e)


#-------------------------  CFX Function  -----------------------

def get_UMs(client, cmd):
    try:
        stdin, stdout, stderr = client.exec_command(cmd)
        out = stdout.readlines()
        error = stderr.read().decode('utf8')
        write_into_file('GET-VM','CFX', out, error, cmd)
        DNs, DUs = [],[]
        for line in out:
            if 'DNs:' in line:
                DNs.extend(re.search(r'\[(.*)\]',line).group(1).split()) # gets DNs Names
            if 'DUs:' in line:
                DUs.extend(re.search(r'\[(.*)\]',line).group(1).split()) # gets DUs Names
        print(f'{len(DNs)} DNs found named: {DNs} \n{len(DUs)} DUs found named: {DUs}')
        
        cscf_flag = False
        DU_flag =False
        nodes = True
        
        VMs, VM, counter = {}, {}, 0
        
        for line in out:
            if 'DU Name:' in line:
                DU_flag=True
            
            if DU_flag and 'Type:' and'CSCF' in line:
                cscf_flag = True
            
            if DU_flag and cscf_flag and 'Node Name:' and 'Common' in line:
                DU_flag=False
                cscf_flag=False
            else: 
                nodes=True
            
            
            if nodes:
                if 'Node Name:' in line:
                    #print(re.search(r'Name:(.*)', line).group(1))
                    VM['Node Name'] = re.search(r'Name:\s+(.*)', line).group(1)
                elif 'Type:' in line:
                    VM['Type'] = re.search(r'Type:\s+(.*)', line).group(1)
                elif 'Release:' in line:
                    VM['Release'] = re.search(r'Release:\s+(.*)', line).group(1)
                elif 'VmName:' in line:
                    VM['VmName'] = re.search(r'VmName:\s+(.*)', line).group(1)
                elif 'Unique ID:' in line:
                    VM['Unique ID'] = re.search(r'Unique ID:\s+(.*)', line).group(1)
                elif 'Node IP:' in line:
                    VM['Node IP'] = re.search(r'Node IP:\s+(.*)', line).group(1)
                    counter += 1
                    VMs[counter] = VM
                    VM = {}
        print(f'Total {counter} VMs found!')
        
        print('VMs an respective IPs are: ')
        output = {}
        for vm_no, vm in VMs.items():
            #print(f'\nVM No: {vm_no}')
            for key, v in vm.items():
                pass
                #print(v)
                #print(key+ ':' + vm[key] )
            #if vm['Type'] not in output.keys():
            output[vm['Node IP']]=vm['Type']
        
        print(output)
        return output
    except Exception as _e:
        print(_e)

def run_on_all_vms(CFX_OAM, vms, cmd):
    """
    with parallelization
    """
    try:
        process = []
        lock = threading.Lock()
        lock.acquire()
        for key, val in vms.items():
            p = threading.Thread(target=parallel_executions, args=(CFX_OAM, cmd, val, key))
            process.append(p)
            p.start()
        for p in process:
            p.join()
        lock.release()
            
    except Exception as _e:
        print(_e)

def run_on_selected(CFX_OAM, vms, vm_list, cmd):
    vm_list = list(map(str.lower,vm_list))
    """
    with parallelization
    """
    process = []
    lock = threading.Lock()
    lock.acquire()
    for key, val in vms.items():
        if val.lower() in vm_list:
            p = threading.Thread(target=parallel_executions, args=(CFX_OAM, cmd, val, key))
            process.append(p)
            p.start()
    for p in process:
        p.join()
    lock.release()
    
    '''
    for key, val in vms.items():
        if val.lower() in vm_list: 
            
            try:
                ip = transport(CFX_OAM, dest_addr = key, local_addr= '192.168.130.50', port=22, user='root', password='yt_xk39b')
                exec_command(ip,cmd)
            except Exception as _e:
                print(_e)
    '''
def get_alarms(CFX_OAM, vms, cmd):
    """
    with parallelization
    """
    
    
    vm_list = ['CIF', 'OAM', 'L2TD']
    vm_list = list(map(str.lower,vm_list))
    def custom_parallel(CFX, cmd, val, key):
        time = datetime.now().strftime('%H_%M')
        try:
            ip = transport(CFX_OAM, dest_addr = key, local_addr= '192.168.130.50', port=22, user='root', password='yt_xk39b')
            _,out,error = ip.exec_command(cmd)
            out = out.read().decode()
            #write_into_file('ALARMS','CFX', out, error, cmd)
            if out:
                with open(f'RtpDump_log_{val}_{key}_{time}_.log', "w") as text_file:
                    text_file.write(out)
                    display_markdown(f'''<span style='background :yellow' > ➜ ***`{val}`*** log are written in filename: ***`RtpDump_{val}_{key}_{time}_.log`***  ✅ </span>''', raw=True)
            else:
                display_markdown(f'''<span style='background :red' > ➜ Failed to fetch logs!</span>''', raw=True)

        except Exception as _e:
            print(_e)
    process = []
    lock = threading.Lock()
    lock.acquire()
    for key, val in vms.items():
        if val.lower() in vm_list:
            p = threading.Thread(target=custom_parallel, args=(CFX_OAM, cmd, val, key))
            process.append(p)
            p.start()
    for p in process:
        p.join()
    lock.release()
    '''
    for key, val in vms.items():
        if val.lower() in vm_list:
            
            try:
                ip = transport(CFX_OAM, dest_addr = key, local_addr= '192.168.130.50', port=22, user='root', password='yt_xk39b')
                _,out,error = ip.exec_command(cmd)
                out = out.read().decode()
                #write_into_file('ALARMS','CFX', out, error, cmd)
                if out:
                    with open(f'RtpDump_log_{val}_{key}_{time}_.log', "w") as text_file:
                        text_file.write(out)
                        print(f'{val} log are written in filename: RtpDump_{val}_{key}_{time}_.log')
                else:
                    print(f'{error.read().decode()}')
            except Exception as _e:
                print(_e)
    '''
def run_all_CSCF(CFX_OAM, vms, cmd):
    time = datetime.now().strftime('%H_%M')
    vm_list = ['CSCF']
    vm_list = list(map(str.lower,vm_list))
    
    """
    with parallelization
    """
    
    def custom_parallel_cscf(CFX, cmd, val, key):
        time = datetime.now().strftime('%H_%M')
        try:
            ip = transport(CFX_OAM, dest_addr = key, local_addr= '192.168.130.50', port=22, user='root', password='yt_xk39b')
            _,out,error = ip.exec_command(cmd)
            out = out.read().decode()
            #write_into_file('ALARMS','CFX', out, error, cmd)
            if out:
                with open(f'{val}_log_{key}_{time}_.log', "w") as text_file:
                    text_file.write(out)
                    display_markdown(f'''<span style='background :yellow' > ➜ ***`{val}`*** log are written in filename: ***`{val}_log_{key}_{time}_.log`***  ✅ </span>''', raw=True)
            else:
                display_markdown(f'''<span style='background :red' > ➜ Failed to fetch logs!</span>''', raw=True)

        except Exception as _e:
            print(_e)
    
    process = []
    lock = threading.Lock()
    lock.acquire()
    for key, val in vms.items():
        if val.lower() in vm_list:
            p = threading.Thread(target=custom_parallel_cscf, args=(CFX_OAM, cmd, val, key))
            process.append(p)
            p.start()
    for p in process:
        p.join()
    lock.release()
    
    
    
    '''
    for key, val in vms.items():
        if val.lower() in vm_list: 
            ip = transport(CFX_OAM, dest_addr = key, local_addr= '192.168.130.50', port=22, user='root', password='yt_xk39b')
            _,out,error = ip.exec_command(cmd)
            out = out.read().decode()
            if out:
                with open(f'{val}_log_{key}_{time}_.log', "w") as text_file:
                    text_file.write(out)
                print(f'{val} log are written in filename: {val}_log_{key}_{time}_.log')
            else:
                print(f'{error.read().decode()}')
    '''
def run_each_type_VM(CFX_OAM, vms, cmd):
    vm_list = []
    
    """
    with parallelization
    """
    try:
        process = []
        lock = threading.Lock()
        lock.acquire()
        for key, val in vms.items():
            if val.lower() not in vm_list:    
                p = threading.Thread(target=parallel_executions, args=(CFX_OAM, cmd, val, key))
                process.append(p)
                p.start()
        for p in process:
            p.join()
        lock.release()
            
    except Exception as _e:
        print(_e)
    '''
    for key, val in vms.items():
        if val.lower() not in vm_list: 
            vm_list.append(val.lower())
            ip = transport(CFX_OAM, dest_addr = key, local_addr= '192.168.130.50', port=22, user='root', password='yt_xk39b')
            exec_command(ip, cmd)
'''



#---------------------- Compatibility Matrix -------------------
def get_ne_version(baseline, target_ne_ver, key=False):
    for k, v in baseline.items():
        if isinstance(v, str) and target_ne_ver.lower() in v.lower():
            return v
    if key:
        ver_list =[key for key, value in baseline.items() if isinstance(value, str) and target_ne_ver.lower() in value.lower()]
    else:
        ver_list =[value for key, value in baseline.items() if isinstance(value, str) and target_ne_ver.lower() in value.lower()]
    return ver_list    

def color_test(sh, target_ver, target_ne, target):
    """
    gets row and col from .xlsx sheet, parse cell location and checks color
    """
    #print(f'target version {target_ver}\n target ne {target_ne}')
    col = re.match("([a-zA-Z]+)([0-9]+)", target_ne).group(1)
    row = re.match("([a-zA-Z]+)([0-9]+)", target_ver).group(2)
    color_in_hex = sh[col+row].fill.start_color.index
    if color_in_hex=='FF00E88E':
        display_markdown(f'''<span style='background :lightgreen' > Compatible `with` ***`{target.get(target_ne)}`*** </span> 😊 ''', raw=True)
    else:
        display_markdown(f'''<span style='background :orange' > Not Compatible `with` ***`{target.get(target_ne)}`*** </span> 😔 ''', raw=True)

def compatibility_matrix(target_ne_ver, ne_versions, sheet='Compa Matrix', path='./Cloud_matrix.xlsx'):
    """
    reads .xlsx file and check compatibility
    """
    if isinstance(path, str) and path.endswith('.xlsx'):
        excel_file = path
        wb = load_workbook(excel_file, data_only = True)
        sh = wb[sheet]

        baseline, target = {}, {}
        for key, *values in sh.iter_rows():
            baseline[key.coordinate] = key.value
            if key.value == ' ': #key.key.coordinate == 'A2': # key.value is None:
                for target_version in values:
                    target[target_version.coordinate] = target_version.value
        
        base  =[value for key, value in baseline.items() if isinstance(value, str) and target_ne_ver.lower() in value.lower()]
        available_target = widgets.Dropdown(options = base) # returns dropdown of listed versions
        
        @interact(Target = available_target) # change1
        def print_version(Target):
            Target = available_target
            for ne in ne_versions:
                target_ne = [key for key, value in target.items() if isinstance(value, str) and ne.lower().strip() == value.lower()]
                target_ver = [key for key, value in baseline.items() if isinstance(value, str) and Target.value.lower()==value.lower()]
                target_ne = ''.join(str(x) for x in target_ne)
                target_ver = ''.join(str(x) for x in target_ver)
                color_test(sh, target_ver, target_ne, target)
    
    else: display_markdown(f'''  Please select ***`.xlsx`*** file 😔''', raw=True)




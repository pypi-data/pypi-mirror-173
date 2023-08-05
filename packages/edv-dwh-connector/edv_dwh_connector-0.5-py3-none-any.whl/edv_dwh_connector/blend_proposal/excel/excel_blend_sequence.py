"""
This module defines Excel blend sequence.
.. since 0.1
"""

# -*- coding: utf-8 -*-
# Copyright (c) 2022 Endeavour Mining
#
# Permission is hereby granted, free of charge, to any person obtaining a copy
# of this software and associated documentation files (the "Software"), to read
# the Software only. Permissions is hereby NOT GRANTED to use, copy, modify,
# merge, publish, distribute, sublicense, and/or sell copies of the Software.
#
# THE SOFTWARE IS PROVIDED "AS IS", WITHOUT WARRANTY OF ANY KIND, EXPRESS OR
# IMPLIED, INCLUDING BUT NOT LIMITED TO THE WARRANTIES OF MERCHANTABILITY,
# FITNESS FOR A PARTICULAR PURPOSE AND NON-INFRINGEMENT. IN NO EVENT SHALL THE
# AUTHORS OR COPYRIGHT HOLDERS BE LIABLE FOR ANY CLAIM, DAMAGES OR OTHER
# LIABILITY, WHETHER IN AN ACTION OF CONTRACT, TORT OR OTHERWISE, ARISING FROM,
# OUT OF OR IN CONNECTION WITH THE SOFTWARE OR THE USE OR OTHER DEALINGS IN THE
# SOFTWARE.

import re
from openpyxl.worksheet.worksheet import Worksheet   # type: ignore
from edv_dwh_connector.blend_proposal.blend_sequence import BlendSequence
from edv_dwh_connector.blend_proposal.excel.start_cell import StartCell
from edv_dwh_connector.blend_proposal.excel.excel_sequence_material \
    import ExcelSequenceMaterial
from edv_dwh_connector.blend_proposal.excel.sequence_material_start_cell \
    import SequenceMaterialStartCell


# pylint: disable=duplicate-code
class ExcelBlendSequence(BlendSequence):
    """
    Excel blend sequence.
    .. since: 0.1
    """

    def __init__(self, sheet: Worksheet, bs_start_cell: StartCell):
        """
        Ctor.
        :param sheet: Worksheet
        :param bs_start_cell: Blend sequence start cell
        """
        self.__sheet = sheet
        self.__bs_start_cell = bs_start_cell

    def name(self) -> str:
        return self.__sheet.cell(
            self.__bs_start_cell.row(), self.__bs_start_cell.column() + 1
        ).value

    def average_grade(self) -> float:
        return self.__value_of("average")

    def soluble_copper(self) -> float:
        return self.__value_of("soluble")

    def as_ppm(self) -> float:
        value = self.__value_of("as")
        if value == 0.0:
            value = self.__value_of("arsenic")
        return value

    def moisture_estimated(self) -> float:
        return self.__value_of("moisture")

    def oxide_transition(self) -> float:
        return self.__value_of("oxide/transition")

    def fresh(self) -> float:
        return self.__value_of("fresh")

    def recovery_blend_estimated(self) -> float:
        return self.__value_of("recovery")

    def materials(self) -> list:
        items = []
        start = self.__bs_start_cell.row() + 1
        for row in range(start, start + self.__total_row()):
            col = self.__bs_start_cell.column()
            name = self.__sheet.cell(row, col).value
            if name is not None and \
                    not re.search("Material", name, re.IGNORECASE):
                items.append(
                    ExcelSequenceMaterial(
                        self.__sheet, self.__bs_start_cell,
                        SequenceMaterialStartCell(row=row, column=col)
                    )
                )
        return items

    def __value_of(self, word) -> float:
        """
        Gets value of key containing word.
        :param word: Word
        :return: Value
        """
        value = 0.0
        rrow = self.__bs_start_cell.row()
        rcol = self.__bs_start_cell.column()
        row = rrow
        count = self.__total_row()
        while row <= rrow + count:
            current = self.__sheet.cell(row, rcol - 3)
            if current.value is not None and \
                    re.search(word, current.value, re.IGNORECASE):
                value = self.__sheet.cell(row, rcol + 6).value
                break
            row = row + 1
        return value

    def __total_row(self) -> int:
        """
        Get total number of rows of a sequence table.
        :return: Number
        """
        number = 0
        row = self.__bs_start_cell.row() + 1
        col = self.__bs_start_cell.column() + 6
        while self.__sheet.cell(row, col).value is not None:
            number = number + 1
            row = row + 1
        return number

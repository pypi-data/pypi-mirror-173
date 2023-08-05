"""
This module defines Excel sequence material.
.. since: 0.1
"""

# -*- coding: utf-8 -*-
# Copyright (c) 2022 Endeavour Mining
#
# Permission is hereby granted, free of charge, to any person obtaining a copy
# of this software and associated documentation files (the "Software"), to read
# the Software only. Permissions is hereby NOT GRANTED to use, copy, modify,
# merge, publish, distribute, sublicense, and/or sell copies of the Software.
#
# THE SOFTWARE IS PROVIDED "AS IS", WITHOUT WARRANTY OF ANY KIND, EXPRESS OR
# IMPLIED, INCLUDING BUT NOT LIMITED TO THE WARRANTIES OF MERCHANTABILITY,
# FITNESS FOR A PARTICULAR PURPOSE AND NON-INFRINGEMENT. IN NO EVENT SHALL THE
# AUTHORS OR COPYRIGHT HOLDERS BE LIABLE FOR ANY CLAIM, DAMAGES OR OTHER
# LIABILITY, WHETHER IN AN ACTION OF CONTRACT, TORT OR OTHERWISE, ARISING FROM,
# OUT OF OR IN CONNECTION WITH THE SOFTWARE OR THE USE OR OTHER DEALINGS IN THE
# SOFTWARE.

import re
from openpyxl.worksheet.worksheet import Worksheet  # type: ignore
from openpyxl.cell import Cell  # type: ignore
from edv_dwh_connector.blend_proposal.blend_material import BlendMaterial
from edv_dwh_connector.blend_proposal.excel.start_cell import StartCell


class ExcelSequenceMaterial(BlendMaterial):
    """
    Excel sequence material.
    .. since: 0.1
    """

    def __init__(
        self, sheet: Worksheet, bs_start_cell: StartCell,
        sm_start_cell: StartCell
    ) -> None:
        """
        Ctor.
        :param sheet: Worksheet
        :param bs_start_cell: Blend sequence start cell
        :param sm_start_cell: Sequence material start cell
        """

        self.__sheet = sheet
        self.__bs_start_cell = bs_start_cell
        self.__sm_start_cell = sm_start_cell

    def machine_type(self) -> str:
        if re.search("SURGE_BIN", self.name(), re.IGNORECASE) or \
                re.search("COS", self.name(), re.IGNORECASE):
            value = "SURGE BIN"
        else:
            value = "CRUSHER"
        return value

    def pit(self) -> str:
        value = ""
        col = self.__pit_column()
        if col != -1:
            value = self.__sheet.cell(self.__sm_start_cell.row(), col).value
        return value

    def name(self) -> str:
        return self.__value_of(0).value

    def au_grade(self) -> float:
        return self.__value_of(1).value

    def sol_cu(self) -> float:
        return self.__value_of(2).value

    def as_ppm(self) -> float:
        return self.__value_of(3).value

    def moisture(self) -> float:
        return self.__value_of(4).value

    def indicative_rec(self) -> float:
        return self.__value_of(5).value

    def bucket(self) -> float:
        return self.__value_of(6).value

    def available_tons(self) -> float:
        return self.__value_of(7).value

    def prop(self) -> float:
        return self.__value_of(8).value

    def __pit_column(self) -> int:
        """
        Gets PIT column.
        :return: Column
        """
        col = -1
        value = self.__sheet.cell(
            self.__bs_start_cell.row() + 1,
            self.__bs_start_cell.column() - 1
        ).value
        if value is not None and re.search("PIT", value, re.IGNORECASE):
            col = self.__bs_start_cell.column() - 1
        return col

    def __value_of(self, pos) -> Cell:
        """
        Gets value at position.
        :param pos: Position
        :return: Cell
        """
        return self.__sheet.cell(
            self.__sm_start_cell.row(), self.__sm_start_cell.column() + pos
        )
